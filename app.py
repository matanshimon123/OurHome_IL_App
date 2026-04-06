import os
import re
import json
import sqlite3
import threading
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from collections import defaultdict
import random
import string
from datetime import datetime, timedelta, timezone
from functools import wraps
from collections import defaultdict
import io
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file
from flask_wtf.csrf import CSRFProtect, generate_csrf
from flask_mail import Mail, Message
import jwt as pyjwt
import secrets
from werkzeug.security import generate_password_hash, check_password_hash
import calendar
from zoneinfo import ZoneInfo
from firebase_config import (firebase_create_user, firebase_verify_login,
                              firebase_send_reset_email, firebase_update_password)

ISRAEL_TZ = ZoneInfo('Asia/Jerusalem')


def now_israel():
    return datetime.now(ISRAEL_TZ).replace(tzinfo=None)


app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'dev-key-change-in-production')
app.permanent_session_lifetime = timedelta(days=30)
app.config['TEMPLATES_AUTO_RELOAD'] = True
DATABASE = os.environ.get('DATABASE_PATH', 'finance_tracker.db')
JWT_SECRET = os.environ.get('JWT_SECRET', app.secret_key)
JWT_EXPIRY_DAYS = 30


def is_admin():
    if hasattr(request, 'api_user'):
        return bool(request.api_user.get('is_admin'))
    return bool(session.get('is_admin'))


csrf = CSRFProtect(app)
app.config['WTF_CSRF_CHECK_DEFAULT'] = False  # We handle CSRF manually below
# Mail config
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = os.environ.get('MAIL_USERNAME', '')
app.config['MAIL_PASSWORD'] = os.environ.get('MAIL_PASSWORD', '')
app.config['MAIL_DEFAULT_SENDER'] = ('OurHome IL', os.environ.get('MAIL_USERNAME', ''))
mail = Mail(app)


# Exempt JSON API endpoints — protected by session auth
@app.after_request
def inject_csrf_token(response):
    response.set_cookie('csrf_token', generate_csrf())
    return response


@app.after_request
def add_no_cache(response):
    if request.endpoint in ('login', 'register'):
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
    return response


@app.after_request
def add_cors_headers(response):
    """Allow mobile app to make API requests"""
    origin = request.headers.get('Origin', '')
    allowed_origins = [
        'capacitor://localhost',
        'http://localhost',
        'https://localhost',
    ]
    if origin in allowed_origins or origin.startswith('http://localhost:'):
        response.headers['Access-Control-Allow-Origin'] = origin
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
        response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'
        response.headers['Access-Control-Allow-Credentials'] = 'true'
    return response


@app.route('/api/<path:path>', methods=['OPTIONS'])
@csrf.exempt
def api_options(path):
    """Handle CORS preflight for all API routes"""
    return '', 204


@app.before_request
def make_session_permanent():
    session.permanent = True


@app.before_request
def check_csrf():
    """Skip CSRF for all API routes, enforce for web forms"""
    if request.method not in ('POST', 'PUT', 'PATCH', 'DELETE'):
        return
    if request.path.startswith('/api/'):
        return  # API routes protected by JWT, not CSRF
    csrf.protect()


@app.before_request
def auto_jwt_auth():
    """Auto-authenticate JWT for API routes and populate session for backward compat"""
    if not request.path.startswith('/api/'):
        return

    token = get_token_from_request()
    if not token:
        return

    payload = decode_jwt_token(token)
    if not payload:
        return

    # Load fresh user data from DB
    try:
        with get_db() as conn:
            user = conn.execute('SELECT * FROM users WHERE id=?', (payload['user_id'],)).fetchone()
            if not user:
                return

        # Set api_user on request
        request.api_user = {
            'user_id': user['id'],
            'username': user['username'],
            'display_name': user['display_name'] or user['username'],
            'family_id': user['family_id'],
            'is_admin': bool(user['is_admin']),
            'email': user['email']
        }

        # Populate session so existing routes work without changes
        session['user_id'] = user['id']
        session['username'] = user['username']
        session['display_name'] = user['display_name'] or user['username']
        session['family_id'] = user['family_id']
        session['is_admin'] = bool(user['is_admin'])
    except Exception:
        pass


def get_db():
    conn = sqlite3.connect(DATABASE, timeout=10)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA journal_mode=WAL')
    conn.execute('PRAGMA synchronous=NORMAL')
    conn.execute('PRAGMA busy_timeout=5000')
    return conn


def generate_invite_code():
    """Generate unique 6-char invite code"""
    with get_db() as conn:
        for _ in range(10):
            code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
            exists = conn.execute('SELECT id FROM families WHERE invite_code=?', (code,)).fetchone()
            if not exists:
                return code
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))  # fallback longer code


def get_family_id():
    if hasattr(request, 'api_user'):
        return request.api_user.get('family_id')
    return session.get('family_id')


# ──────────────────────────────────────────────
# BILLING CYCLE HELPERS
# ──────────────────────────────────────────────
def get_cycle_day(family_id):
    """Get the billing cycle day for a family (default: 1 = calendar month)"""
    if not family_id:
        return 1
    try:
        with get_db() as conn:
            s = conn.execute('SELECT cycle_day FROM family_settings WHERE family_id=?', (family_id,)).fetchone()
            return s['cycle_day'] if s and s['cycle_day'] else 1
    except:
        return 1


def get_cycle_month(family_id):
    """Get current cycle's month string (e.g. '2026-03') based on family's cycle_day.
    
    With cycle_day=10:
      - March 15 → '2026-03' (cycle March = Mar 10 - Apr 9)
      - March 5  → '2026-02' (cycle Feb = Feb 10 - Mar 9)
    With cycle_day=1: same as calendar month (backward compatible)
    """
    now = now_israel()
    cycle_day = get_cycle_day(family_id)

    if cycle_day == 1:
        return now.strftime('%Y-%m')

    if now.day >= cycle_day:
        # We're in current month's cycle
        return now.strftime('%Y-%m')
    else:
        # We're still in previous month's cycle
        prev = now.replace(day=1) - timedelta(days=1)
        return prev.strftime('%Y-%m')


def get_cycle_range(family_id):
    """Get start and end dates of the current billing cycle.
    
    Returns: (start_date, end_date, cycle_label)
    With cycle_day=10, on March 15:
      start = 2026-03-10, end = 2026-04-09, label = 'מרץ 2026 (10.3–9.4)'
    """
    now = now_israel()
    cycle_day = get_cycle_day(family_id)

    hebrew_months = {1:'ינואר',2:'פברואר',3:'מרץ',4:'אפריל',5:'מאי',6:'יוני',
                     7:'יולי',8:'אוגוסט',9:'ספטמבר',10:'אוקטובר',11:'נובמבר',12:'דצמבר'}

    if cycle_day == 1:
        # Standard calendar month
        last_day = calendar.monthrange(now.year, now.month)[1]
        start = now.replace(day=1)
        end = now.replace(day=last_day)
        label = f'{hebrew_months[now.month]} {now.year}'
        return start, end, label

    # Custom cycle day
    if now.day >= cycle_day:
        # Current month's cycle
        cycle_month = now.month
        cycle_year = now.year
    else:
        # Previous month's cycle
        prev = now.replace(day=1) - timedelta(days=1)
        cycle_month = prev.month
        cycle_year = prev.year

    # Start: cycle_day of cycle_month
    try:
        start = datetime(cycle_year, cycle_month, min(cycle_day, calendar.monthrange(cycle_year, cycle_month)[1]))
    except:
        start = datetime(cycle_year, cycle_month, 1)

    # End: day before cycle_day of next month
    if cycle_month == 12:
        next_month, next_year = 1, cycle_year + 1
    else:
        next_month, next_year = cycle_month + 1, cycle_year

    try:
        end = datetime(next_year, next_month, min(cycle_day, calendar.monthrange(next_year, next_month)[1])) - timedelta(days=1)
    except:
        end = datetime(next_year, next_month, 1) - timedelta(days=1)

    label = f'{hebrew_months[cycle_month]} {cycle_year} ({start.day}.{start.month}–{end.day}.{end.month})'
    return start, end, label
def create_jwt_token(user_id, username, display_name, family_id, is_admin=False):
    """Create a JWT token for the user"""
    payload = {
        'user_id': user_id,
        'username': username,
        'display_name': display_name,
        'family_id': family_id,
        'is_admin': is_admin,
        'exp': datetime.utcnow() + timedelta(days=JWT_EXPIRY_DAYS),
        'iat': datetime.utcnow()
    }
    return pyjwt.encode(payload, JWT_SECRET, algorithm='HS256')


def decode_jwt_token(token):
    """Decode and validate a JWT token"""
    try:
        return pyjwt.decode(token, JWT_SECRET, algorithms=['HS256'])
    except pyjwt.ExpiredSignatureError:
        return None
    except pyjwt.InvalidTokenError:
        return None


def get_token_from_request():
    """Extract JWT token from Authorization header"""
    auth_header = request.headers.get('Authorization', '')
    if auth_header.startswith('Bearer '):
        return auth_header[7:]
    return None


def require_api_auth(f):
    """Decorator: require valid JWT token for API endpoints"""
    @wraps(f)
    def decorated(*args, **kwargs):
        token = get_token_from_request()
        if not token:
            return jsonify({'error': 'נדרשת הזדהות', 'code': 'AUTH_REQUIRED'}), 401

        payload = decode_jwt_token(token)
        if not payload:
            return jsonify({'error': 'טוקן לא תקין או פג תוקף', 'code': 'TOKEN_INVALID'}), 401

        # Refresh user data from DB
        with get_db() as conn:
            user = conn.execute('SELECT * FROM users WHERE id=?', (payload['user_id'],)).fetchone()
            if not user:
                return jsonify({'error': 'משתמש לא נמצא', 'code': 'USER_NOT_FOUND'}), 401

        request.api_user = {
            'user_id': user['id'],
            'username': user['username'],
            'display_name': user['display_name'] or user['username'],
            'family_id': user['family_id'],
            'is_admin': bool(user['is_admin']),
            'email': user['email']
        }
        return f(*args, **kwargs)

    return decorated


def require_api_family(f):
    """Decorator: require user to have a family"""
    @wraps(f)
    @require_api_auth
    def decorated(*args, **kwargs):
        if not request.api_user.get('family_id'):
            return jsonify({'error': 'יש להצטרף למשפחה קודם', 'code': 'NO_FAMILY'}), 403
        return f(*args, **kwargs)

    return decorated


def get_api_family_id():
    """Get family_id from API auth or session"""
    if hasattr(request, 'api_user'):
        return request.api_user['family_id']
    return session.get('family_id')


@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        if email:
            firebase_send_reset_email(email)
        # Always show same message — don't reveal if email exists
        flash('אם האימייל קיים במערכת — נשלחה הודעה עם קישור לאיפוס', 'info')
        return redirect(url_for('login'))
    return render_template('forgot_password.html')


@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    with get_db() as conn:
        user = conn.execute(
            'SELECT * FROM users WHERE reset_token=? AND reset_token_exp > ?',
            (token, now_israel().strftime('%Y-%m-%d %H:%M:%S'))
        ).fetchone()
    if not user:
        flash('הקישור לא תקין או פג תוקף', 'error')
        return redirect(url_for('login'))
    if request.method == 'POST':
        pw = request.form.get('password', '')
        pw2 = request.form.get('password2', '')
        if len(pw) < 6:
            flash('סיסמה חייבת להיות לפחות 6 תווים', 'error')
            return render_template('reset_password.html', token=token)
        if pw != pw2:
            flash('הסיסמאות לא תואמות', 'error')
            return render_template('reset_password.html', token=token)
        with get_db() as conn:
            conn.execute(
                'UPDATE users SET password_hash=?, reset_token="", reset_token_exp=NULL WHERE id=?',
                (generate_password_hash(pw), user['id'])
            )
        # Sync to Firebase Auth
        firebase_update_password(user['email'] or user['username'], pw)
        flash('סיסמה שונתה בהצלחה!', 'success')
        return redirect(url_for('login'))
    return render_template('reset_password.html', token=token)


def init_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS families (
            id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL,
            invite_code TEXT UNIQUE NOT NULL, created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT, username TEXT UNIQUE NOT NULL,
            email TEXT DEFAULT '', display_name TEXT DEFAULT '',
            password_hash TEXT NOT NULL, family_id INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (family_id) REFERENCES families(id));
        CREATE TABLE IF NOT EXISTS payments (
            id INTEGER PRIMARY KEY AUTOINCREMENT, family_id INTEGER,
            description TEXT NOT NULL, amount REAL NOT NULL,
            category TEXT DEFAULT 'כללי', date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            month TEXT NOT NULL, year INTEGER NOT NULL, archived BOOLEAN DEFAULT FALSE);
        CREATE TABLE IF NOT EXISTS categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE NOT NULL,
            color TEXT DEFAULT '#007bff');
        CREATE TABLE IF NOT EXISTS archived_cycles (
            id INTEGER PRIMARY KEY AUTOINCREMENT, family_id INTEGER,
            label TEXT NOT NULL, total REAL NOT NULL, count INTEGER NOT NULL,
            archived_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS app_settings (key TEXT PRIMARY KEY, value TEXT NOT NULL);
        CREATE TABLE IF NOT EXISTS shopping_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT, family_id INTEGER NOT NULL,
            name TEXT NOT NULL, quantity INTEGER DEFAULT 1,
            checked BOOLEAN DEFAULT FALSE, image TEXT DEFAULT '',
            category TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS shopping_favorites (
            id INTEGER PRIMARY KEY AUTOINCREMENT, family_id INTEGER NOT NULL,
            name TEXT NOT NULL, quantity INTEGER DEFAULT 1,
            category TEXT DEFAULT '',
            UNIQUE(family_id, name));
        CREATE TABLE IF NOT EXISTS feedings (
            id INTEGER PRIMARY KEY AUTOINCREMENT, family_id INTEGER NOT NULL,
            feeding_type TEXT NOT NULL, amount REAL DEFAULT 0,
            duration INTEGER DEFAULT 0, notes TEXT DEFAULT '',
            date TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS recurring_payments (
            id INTEGER PRIMARY KEY AUTOINCREMENT, family_id INTEGER NOT NULL,
            description TEXT NOT NULL, amount REAL NOT NULL,
            category TEXT DEFAULT 'כללי',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS push_tokens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            token TEXT NOT NULL UNIQUE,
            platform TEXT DEFAULT 'android',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id));
        CREATE TABLE IF NOT EXISTS family_settings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            family_id INTEGER NOT NULL UNIQUE,
            feeding_reminder_hours REAL DEFAULT 0,
            last_feeding_alert TIMESTAMP,
            budget_monthly INTEGER DEFAULT 0,
            budget_daily INTEGER DEFAULT 0,
            budget_alert_80_sent TEXT DEFAULT '',
            budget_alert_100_sent TEXT DEFAULT '',
            FOREIGN KEY (family_id) REFERENCES families(id));
    """)
    for t, c, ct in [
        ('users','email','TEXT DEFAULT ""'),
        ('users','display_name','TEXT DEFAULT ""'),
        ('users','family_id','INTEGER'),
        ('users','is_admin','BOOLEAN DEFAULT FALSE'),
        ('users','reset_token','TEXT DEFAULT ""'),
        ('users','reset_token_exp','TIMESTAMP'),
        ('shopping_items','image','TEXT DEFAULT ""'),
        ('shopping_items','family_id','INTEGER'),
        ('shopping_items','favorite','BOOLEAN DEFAULT FALSE'),
        ('shopping_items','category','TEXT DEFAULT ""'),
        ('payments','family_id','INTEGER'),
        ('feedings','family_id','INTEGER'),
        ('recurring_payments','family_id','INTEGER'),
        ('archived_cycles','family_id','INTEGER'),
        ('archived_cycles','month','TEXT DEFAULT ""'),
        ('family_settings','feeding_reminder_hours','REAL DEFAULT 0'),
        ('family_settings','last_feeding_alert','TIMESTAMP'),
        ('family_settings','budget_monthly','INTEGER DEFAULT 0'),
        ('family_settings','budget_daily','INTEGER DEFAULT 0'),
        ('family_settings','budget_alert_80_sent','TEXT DEFAULT ""'),
        ('family_settings','budget_alert_100_sent','TEXT DEFAULT ""'),
        ('family_settings','cycle_day','INTEGER DEFAULT 1'),
        ('family_settings','last_cycle_archived','TEXT DEFAULT ""'),
        ('categories','family_id','INTEGER DEFAULT NULL'),
        ('users','firebase_uid','TEXT DEFAULT ""'),
    ]:
        try: conn.execute(f'ALTER TABLE {t} ADD COLUMN {c} {ct}')
        except sqlite3.OperationalError: pass
    # ── Performance indexes ──
    for idx in [
        'CREATE INDEX IF NOT EXISTS idx_payments_family_month ON payments(family_id, month, archived)',
        'CREATE INDEX IF NOT EXISTS idx_payments_family_date ON payments(family_id, date)',
        'CREATE INDEX IF NOT EXISTS idx_users_family ON users(family_id)',
        'CREATE INDEX IF NOT EXISTS idx_users_email ON users(email)',
        'CREATE INDEX IF NOT EXISTS idx_users_firebase_uid ON users(firebase_uid)',
        'CREATE INDEX IF NOT EXISTS idx_shopping_family ON shopping_items(family_id, checked)',
        'CREATE INDEX IF NOT EXISTS idx_feedings_family_date ON feedings(family_id, date)',
        'CREATE INDEX IF NOT EXISTS idx_push_tokens_user ON push_tokens(user_id)',
        'CREATE INDEX IF NOT EXISTS idx_recurring_family ON recurring_payments(family_id)',
        'CREATE INDEX IF NOT EXISTS idx_archived_family ON archived_cycles(family_id)',
        'CREATE INDEX IF NOT EXISTS idx_categories_family ON categories(family_id)',
    ]:
        try: conn.execute(idx)
        except: pass
    for cat, color in [
        ('קבועים','#6f42c1'),('משק בית','#28a745'),('קניות - סופר','#ffc107'),
        ('קניות - אופנה','#17a2b8'),('רכב','#dc3545'),('תינוק','#e83e8c'),
        ('בילויים / פנאי','#20c997'),('טיפוח','#fd7e14'),('כללי','#6c757d'),
    ]:
        existing = conn.execute('SELECT id FROM categories WHERE name=? AND family_id IS NULL', (cat,)).fetchone()
        if not existing:
            conn.execute('INSERT INTO categories (name, color, family_id) VALUES (?,?,NULL)', (cat, color))
    conn.commit()
    conn.close()

def require_auth(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        is_api = request.path.startswith('/api/')

        if 'user_id' not in session:
            if is_api:
                return jsonify({'error': 'נדרשת הזדהות', 'code': 'AUTH_REQUIRED'}), 401
            return redirect(url_for('login'))

        # Sync family_id from DB (handles removed members)
        with get_db() as conn:
            user = conn.execute('SELECT family_id FROM users WHERE id=?', (session['user_id'],)).fetchone()
            if user:
                session['family_id'] = user['family_id']
            else:
                session.clear()
                if is_api:
                    return jsonify({'error': 'משתמש לא נמצא', 'code': 'USER_NOT_FOUND'}), 401
                return redirect(url_for('login'))

        if not session.get('family_id') and request.endpoint not in (
                'family_setup', 'create_family', 'join_family', 'logout', 'service_worker',
                'settings', 'api_family_info', 'api_create_family', 'api_join_family',
                'api_get_family_settings', 'api_update_family_settings'):
            if is_api:
                return jsonify({'error': 'יש להצטרף למשפחה קודם', 'code': 'NO_FAMILY'}), 403
            return redirect(url_for('family_setup'))

        return f(*args, **kwargs)

    return decorated


def require_admin(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not is_admin():
            if request.path.startswith('/api/'):
                return jsonify({'error': 'נדרשות הרשאות מנהל', 'code': 'ADMIN_REQUIRED'}), 403
            return redirect(url_for('home'))
        return f(*args, **kwargs)

    return decorated


@app.route('/admin')
@require_admin
def admin_dashboard():
    with get_db() as conn:
        # ── KPIs ──
        total_users = conn.execute('SELECT COUNT(*) as c FROM users').fetchone()['c']
        total_families = conn.execute('SELECT COUNT(*) as c FROM families').fetchone()['c']
        # משפחות עם 2+ חברים = פעילות אמיתית
        active_families = conn.execute('''
            SELECT COUNT(*) as c FROM (
                SELECT family_id FROM users WHERE family_id IS NOT NULL
                GROUP BY family_id HAVING COUNT(*) >= 2
            )''').fetchone()['c']

        week_ago = (now_israel() - timedelta(days=7)).strftime('%Y-%m-%d')
        month_ago = (now_israel() - timedelta(days=30)).strftime('%Y-%m-%d')

        new_week = conn.execute('SELECT COUNT(*) as c FROM users WHERE created_at >= ?', (week_ago,)).fetchone()['c']
        new_month = conn.execute('SELECT COUNT(*) as c FROM users WHERE created_at >= ?', (month_ago,)).fetchone()['c']

        # ── גרף הרשמות יומי — 30 ימים אחרונים ──
        reg_daily = conn.execute('''
            SELECT DATE(created_at) as day, COUNT(*) as cnt
            FROM users
            WHERE created_at >= ?
            GROUP BY DATE(created_at)
            ORDER BY day
        ''', (month_ago,)).fetchall()

        # ── גרף הרשמות חודשי — 12 חודשים אחרונים ──
        reg_monthly = conn.execute('''
            SELECT strftime('%Y-%m', created_at) as month, COUNT(*) as cnt
            FROM users
            GROUP BY strftime('%Y-%m', created_at)
            ORDER BY month DESC
            LIMIT 12
        ''').fetchall()
        reg_monthly = list(reversed(reg_monthly))

        # ── פעילות תשלומים — ללא פרטים ──
        payments_month = conn.execute('''
            SELECT COUNT(*) as c FROM payments
            WHERE date >= ?
        ''', (month_ago,)).fetchone()['c']

        payments_by_month = conn.execute('''
            SELECT month, COUNT(*) as cnt, COUNT(DISTINCT family_id) as families
            FROM payments
            GROUP BY month
            ORDER BY month DESC
            LIMIT 12
        ''').fetchall()
        payments_by_month = list(reversed(payments_by_month))

        # ── משפחות לפי גודל ──
        family_sizes = conn.execute('''
            SELECT COUNT(*) as members, COUNT(*) as families
            FROM users
            WHERE family_id IS NOT NULL
            GROUP BY family_id
        ''').fetchall()
        solo = sum(1 for f in family_sizes if f['members'] == 1)
        pairs = sum(1 for f in family_sizes if f['members'] == 2)
        large = sum(1 for f in family_sizes if f['members'] >= 3)

        # ── משתמשים ללא משפחה ──
        no_family = conn.execute(
            'SELECT COUNT(*) as c FROM users WHERE family_id IS NULL'
        ).fetchone()['c']

    return render_template('admin.html',
                           total_users=total_users,
                           total_families=total_families,
                           active_families=active_families,
                           new_week=new_week,
                           new_month=new_month,
                           no_family=no_family,
                           solo=solo, pairs=pairs, large=large,
                           payments_month=payments_month,
                           reg_daily=[dict(r) for r in reg_daily],
                           reg_monthly=[dict(r) for r in reg_monthly],
                           payments_by_month=[dict(r) for r in payments_by_month],
                           now_israel=now_israel,
                           )


@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session and session.get('family_id'):
        return redirect(url_for('home'))
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        password = request.form['password']

        # Verify with Firebase Auth
        fb_uid, error = firebase_verify_login(email, password)
        if error:
            flash(error, 'error')
            return render_template('login.html')

        with get_db() as conn:
            # Find user by email or firebase_uid
            user = conn.execute('SELECT * FROM users WHERE email=? OR firebase_uid=?',
                                (email, fb_uid or '')).fetchone()
            if not user:
                flash('משתמש לא נמצא. יש להירשם קודם', 'error')
                return render_template('login.html')
            # Update firebase_uid if not set
            if fb_uid and not user['firebase_uid']:
                conn.execute('UPDATE users SET firebase_uid=? WHERE id=?', (fb_uid, user['id']))
            session.clear()
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['display_name'] = user['display_name'] or user['username']
            session['family_id'] = user['family_id']
            session['is_admin'] = bool(user['is_admin'])
            if not user['family_id']: return redirect(url_for('family_setup'))
            return redirect(url_for('home'))
    return render_template('login.html')


@app.route('/register', methods=['GET', 'POST'])
def register():
    if 'user_id' in session and session.get('family_id'):
        return redirect(url_for('home'))
    if request.method == 'POST':
        dn = request.form['display_name'].strip()
        un = request.form['username'].strip().lower()
        em = request.form.get('email', '').strip()
        pw = request.form['password']
        pw2 = request.form['password2']
        import re
        if not re.match(r'^[a-zA-Z0-9_]{3,20}$', un):
            flash('שם משתמש חייב להכיל 3-20 תווים באנגלית, מספרים או _', 'error')
            return render_template('register.html')
        if not em or not re.match(r'^[^@]+@[^@]+\.[^@]+$', em):
            flash('כתובת אימייל נדרשת ותקינה', 'error')
            return render_template('register.html')
        if len(pw) < 6:
            flash('סיסמה חייבת להיות לפחות 6 תווים', 'error')
            return render_template('register.html')
        if pw != pw2:
            flash('הסיסמאות לא תואמות', 'error')
            return render_template('register.html')

        # Create in Firebase Auth first
        fb_uid, fb_error = firebase_create_user(em, pw, dn)
        if fb_error:
            flash(fb_error, 'error')
            return render_template('register.html')

        with get_db() as conn:
            if conn.execute('SELECT id FROM users WHERE username=?', (un,)).fetchone():
                flash('שם משתמש כבר תפוס', 'error')
                return render_template('register.html')
            if conn.execute('SELECT id FROM users WHERE email=?', (em,)).fetchone():
                flash('כתובת אימייל כבר רשומה', 'error')
                return render_template('register.html')
            conn.execute('INSERT INTO users (username,email,display_name,password_hash,firebase_uid) VALUES (?,?,?,?,?)',
                         (un, em, dn, generate_password_hash(pw), fb_uid or ''))
            user = conn.execute('SELECT * FROM users WHERE username=?', (un,)).fetchone()
            session.clear()
            session['user_id'] = user['id']
            session['username'] = un
            session['display_name'] = dn
            flash('נרשמת בהצלחה!', 'success')
            return redirect(url_for('family_setup'))
    return render_template('register.html')


@app.route('/logout')
def logout():
    session.clear()
    response = redirect(url_for('login'))
    response.delete_cookie('csrf_token')
    return response


@app.route('/change-password', methods=['POST'])
@require_auth
def change_password():
    current_pw = request.form.get('current_password', '')
    new_pw = request.form.get('new_password', '')
    if len(new_pw) < 6:
        flash('סיסמה חדשה חייבת להיות לפחות 6 תווים', 'error')
        return redirect(url_for('settings'))

    with get_db() as conn:
        user = conn.execute('SELECT * FROM users WHERE id=?', (session['user_id'],)).fetchone()
        if user and check_password_hash(user['password_hash'], current_pw):
            conn.execute('UPDATE users SET password_hash=? WHERE id=?',
                         (generate_password_hash(new_pw), session['user_id']))
            # Sync to Firebase Auth
            firebase_update_password(user['email'] or user['username'], new_pw)
            # שליחת מייל בתוך ה-with כשה-conn עדיין פתוח
            try:
                if user['email']:
                    msg = Message(
                        subject='סיסמתך שונתה — OurHome IL',
                        recipients=[user['email']],
                        html=f'''
                        <div dir="rtl" style="font-family:Arial;max-width:500px;margin:0 auto;">
                            <h2>סיסמה שונתה</h2>
                            <p>הסיסמה לחשבון שלך באפליקציית OurHome IL שונתה זה עתה.</p>
                            <p style="color:#888;font-size:0.85rem;">
                                אם לא ביקשת שינוי זה — פנה אלינו מיד.
                            </p>
                        </div>
                        '''
                    )
                    mail.send(msg)
            except Exception as e:
                print(f'Mail error: {e}')
            flash('סיסמה שונתה בהצלחה!', 'success')
            return redirect(url_for('home'))

    flash('סיסמה נוכחית שגויה', 'error')
    return redirect(url_for('settings'))


@app.route('/family')
@require_auth
def family_setup():
    family = None
    members = []
    if session.get('family_id'):
        with get_db() as conn:
            family = conn.execute('SELECT * FROM families WHERE id=?', (session['family_id'],)).fetchone()
            members = conn.execute('SELECT id,display_name,username FROM users WHERE family_id=?',
                                   (session['family_id'],)).fetchall()
    return render_template('family_setup.html', family=family, members=members)


@app.route('/family/create', methods=['POST'])
def create_family():
    if 'user_id' not in session: return redirect(url_for('login'))
    name = request.form['family_name'].strip()
    code = generate_invite_code()
    with get_db() as conn:
        conn.execute('INSERT INTO families (name,invite_code,created_by) VALUES (?,?,?)',
                     (name, code, session['user_id']))
        fam = conn.execute('SELECT * FROM families WHERE invite_code=?', (code,)).fetchone()
        conn.execute('UPDATE users SET family_id=? WHERE id=?', (fam['id'], session['user_id']))
        session['family_id'] = fam['id']
    flash(f'משפחה "{name}" נוצרה! קוד: {code}', 'success')
    return redirect(url_for('family_setup'))


@app.route('/family/join', methods=['POST'])
def join_family():
    if 'user_id' not in session: return redirect(url_for('login'))
    code = request.form['invite_code'].strip().upper()
    with get_db() as conn:
        fam = conn.execute('SELECT * FROM families WHERE invite_code=?', (code,)).fetchone()
        if not fam:
            flash('קוד הזמנה לא נמצא', 'error')
            return redirect(url_for('family_setup'))
        conn.execute('UPDATE users SET family_id=? WHERE id=?', (fam['id'], session['user_id']))
        session['family_id'] = fam['id']
        send_push_to_family(fam['id'], '👨‍👩‍👧 חבר חדש במשפחה!',
                            f'{session.get("display_name", "")} הצטרף/ה למשפחה',
                            exclude_user_id=session['user_id'])
    flash(f'הצטרפת למשפחת {fam["name"]}!', 'success')
    return redirect(url_for('family_setup'))


@app.route('/sw.js')
def service_worker():
    return send_file('static/sw.js', mimetype='application/javascript')


@app.route('/')
def index():
    if 'user_id' in session: return redirect(url_for('home'))
    return redirect(url_for('login'))


@app.route('/home')
@require_auth
def home():
    return render_template('home.html')


@app.route('/api/home-summary')
@require_auth
def home_summary():
    fid = get_family_id()
    now = now_israel()
    cm = get_cycle_month(fid)
    today = now.strftime('%Y-%m-%d')
    with get_db() as conn:
        # Finance
        fin = conn.execute(
            'SELECT COALESCE(SUM(amount),0) as total, COUNT(*) as count FROM payments WHERE month=? AND archived=FALSE AND family_id=?',
            (cm, fid)).fetchone()
        last_payment = conn.execute(
            'SELECT description, amount FROM payments WHERE month=? AND archived=FALSE AND family_id=? ORDER BY date DESC LIMIT 1',
            (cm, fid)).fetchone()
        # Shopping
        shop_total = conn.execute('SELECT COUNT(*) as total FROM shopping_items WHERE family_id=?', (fid,)).fetchone()[
            'total']
        shop_done = conn.execute('SELECT COUNT(*) as done FROM shopping_items WHERE family_id=? AND checked=TRUE',
                                 (fid,)).fetchone()['done']
        shop_left = shop_total - shop_done
        # Baby
        baby_bottles = conn.execute(
            'SELECT COUNT(*) as c, COALESCE(SUM(amount),0) as ml FROM feedings WHERE family_id=? AND date(date)=? AND feeding_type=?',
            (fid, today, 'bottle')).fetchone()
        baby_bf = conn.execute(
            'SELECT COUNT(*) as c FROM feedings WHERE family_id=? AND date(date)=? AND feeding_type=?',
            (fid, today, 'breastfeeding')).fetchone()
        baby_diapers = conn.execute(
            'SELECT COUNT(*) as c FROM feedings WHERE family_id=? AND date(date)=? AND feeding_type=?',
            (fid, today, 'diaper')).fetchone()
        baby_sleep = conn.execute(
            'SELECT COUNT(*) as c, COALESCE(SUM(amount),0) as mins FROM feedings WHERE family_id=? AND date(date)=? AND feeding_type=?',
            (fid, today, 'sleep')).fetchone()
        baby_all = conn.execute('SELECT COUNT(*) as c FROM feedings WHERE family_id=? AND date(date)=?',
                                (fid, today)).fetchone()
        last_feed = conn.execute(
            'SELECT date FROM feedings WHERE family_id=? AND date(date)=? AND feeding_type IN (?,?,?) ORDER BY time(date) DESC LIMIT 1',
            (fid, today, 'bottle', 'breastfeeding', 'solid')).fetchone()
        last_feed_ago = '--'
        if last_feed:
            try:
                ldt = datetime.strptime(last_feed['date'].split('.')[0], '%Y-%m-%d %H:%M:%S')
                mins = int((now - ldt).total_seconds() // 60)
                if mins < 0:
                    last_feed_ago = 'עוד מעט'
                elif mins < 60:
                    last_feed_ago = f'{mins} דק\''
                else:
                    last_feed_ago = f'{mins // 60}:{mins % 60:02d}h'
            except:
                pass
        sleep_mins = int(baby_sleep['mins']) if baby_sleep['mins'] else 0
        sleep_str = f'{sleep_mins // 60}:{sleep_mins % 60:02d}h' if sleep_mins >= 60 else (
            f'{sleep_mins} דק\'' if sleep_mins > 0 else '--')
    start, end, cycle_label = get_cycle_range(fid)
    return jsonify({
        'finance': {'total': float(fin['total']), 'count': fin['count'],
                    'last': {'desc': last_payment['description'],
                             'amount': float(last_payment['amount'])} if last_payment else None,
                    'cycle_label': cycle_label},
        'shopping': {'total': shop_total, 'done': shop_done, 'left': shop_left},
        'baby': {
            'count': baby_all['c'],
            'total_ml': float(baby_bottles['ml']),
            'bottles': baby_bottles['c'],
            'breastfeedings': baby_bf['c'],
            'diapers': baby_diapers['c'],
            'sleep_str': sleep_str,
            'last_ago': last_feed_ago
        }
    })


@app.route('/settings')
@require_auth
def settings():
    family = None
    members = []
    is_family_admin = False
    if session.get('family_id'):
        with get_db() as conn:
            family = conn.execute('SELECT * FROM families WHERE id=?', (session['family_id'],)).fetchone()
            members = conn.execute('SELECT id,display_name,username FROM users WHERE family_id=?',
                                   (session['family_id'],)).fetchall()
            if family and family['created_by'] == session['user_id']:
                is_family_admin = True
    return render_template('settings.html', family=family, members=members, is_family_admin=is_family_admin)


@app.route('/update-profile', methods=['POST'])
@require_auth
def update_profile():
    dn = request.form.get('display_name', '').strip()
    if dn:
        with get_db() as conn:
            conn.execute('UPDATE users SET display_name=? WHERE id=?', (dn, session['user_id']))
        session['display_name'] = dn
        flash('שם תצוגה עודכן!', 'success')
    return redirect(url_for('settings'))


@app.route('/api/categories', methods=['GET'])
@require_auth
def get_categories():
    fid = get_family_id()
    with get_db() as conn:
        cats = conn.execute(
            'SELECT id, name, color, family_id FROM categories '
            'WHERE family_id IS NULL OR family_id=? ORDER BY name',
            (fid,)).fetchall()
    return jsonify([{'id': c['id'], 'name': c['name'], 'color': c['color'],
                     'is_default': c['family_id'] is None} for c in cats])


@app.route('/api/categories', methods=['POST'])
@require_auth
def add_category():
    data = request.get_json()
    name = data.get('name', '').strip()
    color = data.get('color', '#6c757d')
    if not name: return jsonify({'error': 'Missing name'}), 400
    fid = get_family_id()
    with get_db() as conn:
        existing = conn.execute(
            'SELECT id FROM categories WHERE name=? AND (family_id IS NULL OR family_id=?)',
            (name, fid)).fetchone()
        if existing:
            return jsonify({'error': 'קטגוריה כבר קיימת'}), 400
        conn.execute('INSERT INTO categories (name, color, family_id) VALUES (?,?,?)', (name, color, fid))
    return jsonify({'success': True}), 201


@app.route('/api/categories/<int:cat_id>', methods=['DELETE'])
@csrf.exempt
@require_auth
def delete_category(cat_id):
    fid = get_family_id()
    with get_db() as conn:
        cat = conn.execute('SELECT * FROM categories WHERE id=?', (cat_id,)).fetchone()
        if not cat:
            return jsonify({'error': 'קטגוריה לא נמצאה'}), 404
        if cat['family_id'] is None:
            return jsonify({'error': 'לא ניתן למחוק קטגוריה ברירת מחדל'}), 403
        if cat['family_id'] != fid:
            return jsonify({'error': 'אין הרשאה'}), 403
        conn.execute('DELETE FROM categories WHERE id=?', (cat_id,))
    return jsonify({'success': True})


@app.route('/api/family/remove-member', methods=['POST'])
@require_auth
def remove_family_member():
    fid = get_family_id()
    data = request.get_json() or {}
    target_id = data.get('user_id')
    current_user_id = int(request.api_user['user_id'] if hasattr(request, 'api_user') else session.get('user_id'))
    with get_db() as conn:
        family = conn.execute('SELECT created_by FROM families WHERE id=?', (fid,)).fetchone()
        if not family or int(family['created_by']) != current_user_id:
            return jsonify({'error': 'Only admin can remove members'}), 403
        if not target_id:
            return jsonify({'error': 'user_id נדרש'}), 400
        target_id = int(target_id)
        if target_id == current_user_id:
            return jsonify({'error': 'Cannot remove yourself'}), 400
        conn.execute('UPDATE users SET family_id=NULL WHERE id=? AND family_id=?', (target_id, fid))
    return jsonify({'success': True})


@app.route('/api/family/leave', methods=['POST'])
@require_auth
def leave_family():
    fid = get_family_id()
    with get_db() as conn:
        family = conn.execute('SELECT created_by FROM families WHERE id=?', (fid,)).fetchone()
        if family and family['created_by'] == session['user_id']:
            return jsonify({'error': 'Admin cannot leave. Delete family or transfer ownership first.'}), 400
        conn.execute('UPDATE users SET family_id=NULL WHERE id=?', (session['user_id'],))
    session['family_id'] = None
    return jsonify({'success': True})


@app.route('/dashboard')
@require_auth
def dashboard():
    fid = get_family_id()
    cm = get_cycle_month(fid)
    cy = int(cm.split('-')[0])
    now = now_israel()
    today_str = now.strftime('%Y-%m-%d')
    # Week start = Sunday (Israel work week)
    week_start = (now - timedelta(days=now.weekday() + 1 if now.weekday() != 6 else 0)).strftime('%Y-%m-%d')
    with get_db() as conn:
        payments = conn.execute(
            'SELECT * FROM payments WHERE month=? AND archived=FALSE AND family_id=? ORDER BY date DESC',
            (cm, fid)).fetchall()
        mt = conn.execute(
            'SELECT COALESCE(SUM(amount),0) as total FROM payments WHERE month=? AND archived=FALSE AND family_id=?',
            (cm, fid)).fetchone()['total']
        today_total = conn.execute(
            'SELECT COALESCE(SUM(amount),0) as total FROM payments WHERE month=? AND archived=FALSE AND family_id=? AND date(date)=?',
            (cm, fid, today_str)).fetchone()['total']
        week_total = conn.execute(
            'SELECT COALESCE(SUM(amount),0) as total FROM payments WHERE month=? AND archived=FALSE AND family_id=? AND date(date)>=?',
            (cm, fid, week_start)).fetchone()['total']
        cs = conn.execute(
            'SELECT p.category,SUM(p.amount) as total,c.color FROM payments p LEFT JOIN categories c ON p.category=c.name AND (c.family_id IS NULL OR c.family_id=p.family_id) WHERE p.month=? AND p.archived=FALSE AND p.family_id=? GROUP BY p.category ORDER BY total DESC',
            (cm, fid)).fetchall()
        lar = conn.execute("SELECT value FROM app_settings WHERE key=?", (f'last_archived_total_{fid}',)).fetchone()
        pmt = float(lar['value']) if lar else \
            conn.execute('SELECT COALESCE(SUM(amount),0) as total FROM payments WHERE month=? AND family_id=?',
                         ((now.replace(day=1) - timedelta(days=1)).strftime('%Y-%m'), fid)).fetchone()['total']
        cats = conn.execute(
            'SELECT name FROM categories WHERE family_id IS NULL OR family_id=? ORDER BY name',
            (fid,)).fetchall()
        # Check if family needs onboarding (no family_settings row or cycle_day not set)
        fs = conn.execute('SELECT cycle_day, budget_monthly, budget_daily FROM family_settings WHERE family_id=?', (fid,)).fetchone()
        needs_onboarding = not fs
    start, end, cycle_label = get_cycle_range(fid)
    days_elapsed = (now.date() - (start.date() if isinstance(start, datetime) else start)).days + 1
    da = mt / days_elapsed if days_elapsed > 0 else 0
    cycle_day = get_cycle_day(fid)
    budget_monthly = fs['budget_monthly'] if fs and fs['budget_monthly'] else 0
    budget_daily = fs['budget_daily'] if fs and fs['budget_daily'] else 0
    return render_template('dashboard.html', payments=payments, monthly_total=mt, daily_average=da,
                           weekly_average=da * 7, today_total=today_total, week_total=week_total,
                           category_stats=cs, prev_month_total=pmt,
                           comparison_pct=((mt - pmt) / pmt * 100) if pmt > 0 else 0,
                           categories=cats, current_month=cm, comparison_label='vs Last Month',
                           cycle_label=cycle_label, cycle_day=cycle_day,
                           cycle_start=start.strftime('%Y-%m-%d'), cycle_end=end.strftime('%Y-%m-%d'),
                           needs_onboarding=needs_onboarding,
                           budget_monthly=budget_monthly, budget_daily=budget_daily)


@app.route('/add_payment', methods=['POST'])
@require_auth
def add_payment():
    fid = get_family_id()
    cm = get_cycle_month(fid)
    cy = int(cm.split('-')[0])
    try:
        amount = float(request.form.get('amount', 0))
        if amount <= 0:
            raise ValueError
    except (ValueError, TypeError):
        flash('סכום לא תקין', 'error')
        return redirect(url_for('dashboard'))
    desc = request.form.get('description', '').strip()
    if not desc:
        flash('נא להכניס תיאור', 'error')
        return redirect(url_for('dashboard'))
    with get_db() as conn:
        conn.execute('INSERT INTO payments (family_id,description,amount,category,month,year,date) VALUES (?,?,?,?,?,?,?)',
                     (fid, desc, amount, request.form.get('category', 'כללי'), cm, cy,
                      now_israel().strftime('%Y-%m-%d %H:%M:%S')))

    # Push notification + budget check
    user_id = session.get('user_id')
    user_name = session.get('display_name', '')
    send_push_to_family(fid,
        '💰 הוצאה חדשה',
        f'{user_name} הוסיף/ה: {desc} — ₪{amount:.0f}',
        exclude_user_id=user_id)
    check_budget_alerts(fid)

    return redirect(url_for('dashboard'))


@app.route('/api/payments/add', methods=['POST'])
@require_auth
def add_payment_api():
    fid = get_family_id()
    data = request.get_json()
    cm = get_cycle_month(fid)
    cy = int(cm.split('-')[0])
    desc = data.get('description', '').strip()
    amount = data.get('amount', 0)
    if not desc or amount is None or float(amount) <= 0:
        return jsonify({'error': 'Invalid data'}), 400
    with get_db() as conn:
        conn.execute('INSERT INTO payments (family_id,description,amount,category,month,year,date) VALUES (?,?,?,?,?,?,?)',
                     (fid, data['description'], data['amount'], data.get('category', 'כללי'), cm, cy,
                      now_israel().strftime('%Y-%m-%d %H:%M:%S')))

    # Push notification to family
    user_id = request.api_user['user_id'] if hasattr(request, 'api_user') else session.get('user_id')
    user_name = request.api_user['display_name'] if hasattr(request, 'api_user') else session.get('display_name', '')
    send_push_to_family(fid,
        '💰 הוצאה חדשה',
        f'{user_name} הוסיף/ה: {desc} — ₪{float(amount):.0f}',
        exclude_user_id=user_id)

    # Check budget thresholds
    check_budget_alerts(fid)

    return jsonify({'success': True}), 201


@app.route('/api/payments', methods=['GET'])
@require_auth
def get_payments():
    fid = get_family_id()
    cm = get_cycle_month(fid)
    with get_db() as conn:
        ps = conn.execute(
            'SELECT p.id,p.description,p.amount,p.category,p.date,COALESCE(c.color,\'#6c757d\') as color FROM payments p LEFT JOIN categories c ON p.category=c.name AND (c.family_id IS NULL OR c.family_id=p.family_id) WHERE p.month=? AND p.archived=FALSE AND p.family_id=? ORDER BY p.date DESC',
            (cm, fid)).fetchall()
    return jsonify([{'id': p['id'], 'description': p['description'], 'amount': float(p['amount']),
                     'category': p['category'], 'color': p['color'],
                     'date': p['date'].split(' ')[0] if p['date'] else ''} for p in ps])


@app.route('/api/payments/<int:pid>', methods=['PUT'])
@require_auth
def update_payment(pid):
    fid = get_family_id()
    data = request.get_json()
    allowed = ['description', 'amount', 'category']
    with get_db() as conn:
        for f in allowed:
            if f in data: conn.execute(f'UPDATE payments SET {f}=? WHERE id=? AND family_id=?', (data[f], pid, fid))
        # Get updated payment info for push
        p = conn.execute('SELECT description, amount FROM payments WHERE id=? AND family_id=?', (pid, fid)).fetchone()

    # Push notification
    if p:
        user_id = request.api_user['user_id'] if hasattr(request, 'api_user') else session.get('user_id')
        user_name = request.api_user['display_name'] if hasattr(request, 'api_user') else session.get('display_name', '')
        send_push_to_family(fid,
            '✏️ תשלום עודכן',
            f'{user_name} עדכן/ה: {p["description"]} — ₪{float(p["amount"]):.0f}',
            exclude_user_id=user_id)
        check_budget_alerts(fid)

    return jsonify({'success': True})


@app.route('/delete_payment/<int:pid>', methods=['POST', 'GET'])
@require_auth
def delete_payment(pid):
    fid = get_family_id()
    with get_db() as conn:
        p = conn.execute('SELECT description, amount FROM payments WHERE id=? AND family_id=?', (pid, fid)).fetchone()
        conn.execute('DELETE FROM payments WHERE id=? AND family_id=?', (pid, fid))

    if p:
        user_id = request.api_user['user_id'] if hasattr(request, 'api_user') else session.get('user_id')
        user_name = request.api_user['display_name'] if hasattr(request, 'api_user') else session.get('display_name', '')
        send_push_to_family(fid,
            '🗑️ תשלום נמחק',
            f'{user_name} מחק/ה: {p["description"]} — ₪{float(p["amount"]):.0f}',
            exclude_user_id=user_id)

    if request.is_json or request.args.get('api') or request.method == 'POST':
        return jsonify({'success': True})
    return redirect(url_for('dashboard'))


@app.route('/archive_month', methods=['POST'])
@require_auth
def archive_month():
    fid = get_family_id()
    cm = get_cycle_month(fid)
    start, end, cycle_label = get_cycle_range(fid)
    with get_db() as conn:
        existing = conn.execute('SELECT id FROM archived_cycles WHERE family_id=? AND label=?',
                                (fid, cycle_label)).fetchone()
        if existing:
            flash('מחזור זה כבר מאורכב', 'info')
            return redirect(url_for('dashboard'))
        r = conn.execute(
            'SELECT COALESCE(SUM(amount),0) as total,COUNT(*) as count FROM payments WHERE month=? AND archived=FALSE AND family_id=?',
            (cm, fid)).fetchone()
        if r['count'] > 0:
            conn.execute('INSERT INTO archived_cycles (family_id,label,total,count,month) VALUES (?,?,?,?,?)',
                         (fid, cycle_label, r['total'], r['count'], cm))
            conn.execute('INSERT OR REPLACE INTO app_settings (key,value) VALUES (?,?)',
                         (f'last_archived_total_{fid}', str(r['total'])))
            conn.execute('UPDATE payments SET archived=TRUE WHERE month=? AND archived=FALSE AND family_id=?',
                         (cm, fid))
            flash(f'ארכוב: {r["count"]} תשלומים, ₪{r["total"]:.2f}', 'success')
        else:
            flash('אין תשלומים לארכוב', 'info')
    return redirect(url_for('dashboard'))


@app.route('/export_csv')
@require_auth
def export_csv():
    fid = get_family_id()
    cm = get_cycle_month(fid)
    now = now_israel()
    with get_db() as conn:
        ps = conn.execute(
            'SELECT description,amount,category,date FROM payments WHERE month=? AND family_id=? ORDER BY date DESC',
            (cm, fid)).fetchall()
        cats = conn.execute(
            'SELECT category,SUM(amount) as total FROM payments WHERE month=? AND family_id=? GROUP BY category ORDER BY total DESC',
            (cm, fid)).fetchall()
        # נתונים יומיים
        daily_raw = conn.execute(
            'SELECT date,amount FROM payments WHERE month=? AND family_id=?',
            (cm, fid)).fetchall()

    # חישובים
    total = sum(float(p['amount']) for p in ps)
    avg = total / len(ps) if ps else 0
    max_p = max(ps, key=lambda x: x['amount'], default=None)
    days_in_month = calendar.monthrange(now.year, now.month)[1]
    daily_avg = total / now.day if now.day > 0 else 0

    daily_map = defaultdict(float)
    for r in daily_raw:
        try:
            day = datetime.strptime(r['date'].split('.')[0], '%Y-%m-%d %H:%M:%S').day
            daily_map[day] += float(r['amount'])
        except:
            pass

    month_heb = ['ינואר', 'פברואר', 'מרץ', 'אפריל', 'מאי', 'יוני',
                 'יולי', 'אוגוסט', 'ספטמבר', 'אוקטובר', 'נובמבר', 'דצמבר']
    month_name = month_heb[now.month - 1]

    wb = Workbook()

    # ── גיליון 1: סיכום ── #
    ws1 = wb.active
    ws1.title = 'סיכום'
    ws1.sheet_view.rightToLeft = True
    ws1.column_dimensions['A'].width = 22
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 22
    ws1.column_dimensions['D'].width = 18

    def hdr_cell(ws, row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=True, color='FFFFFF', size=11)
        c.fill = PatternFill('solid', fgColor='1E3A5F')
        c.alignment = Alignment(horizontal='center', vertical='center')
        return c

    def stat_label(ws, row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=True, size=10, color='374151')
        c.alignment = Alignment(horizontal='right')
        return c

    def stat_val(ws, row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=True, size=13, color='2563EB')
        c.alignment = Alignment(horizontal='center')
        return c

    # כותרת ראשית
    ws1.merge_cells('A1:D1')
    t = ws1['A1']
    t.value = f'דוח הוצאות — {month_name} {now.year}'
    t.font = Font(bold=True, size=14, color='FFFFFF')
    t.fill = PatternFill('solid', fgColor='2563EB')
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 30

    # כרטיסי סטטיסטיקה
    ws1.merge_cells('A3:B3')
    hdr_cell(ws1, 3, 1, 'סה"כ הוצאות')
    ws1.merge_cells('C3:D3')
    hdr_cell(ws1, 3, 3, 'מספר תשלומים')
    ws1.merge_cells('A4:B4')
    stat_val(ws1, 4, 1, f'₪{total:,.0f}')
    ws1.merge_cells('C4:D4')
    stat_val(ws1, 4, 3, len(ps))
    ws1.row_dimensions[4].height = 25

    ws1.merge_cells('A6:B6')
    hdr_cell(ws1, 6, 1, 'ממוצע לתשלום')
    ws1.merge_cells('C6:D6')
    hdr_cell(ws1, 6, 3, 'ממוצע יומי')
    ws1.merge_cells('A7:B7')
    stat_val(ws1, 7, 1, f'₪{avg:,.0f}')
    ws1.merge_cells('C7:D7')
    stat_val(ws1, 7, 3, f'₪{daily_avg:,.0f}')
    ws1.row_dimensions[7].height = 25

    if max_p:
        ws1.merge_cells('A9:B9')
        hdr_cell(ws1, 9, 1, 'הוצאה הגדולה ביותר')
        ws1.merge_cells('C9:D9')
        hdr_cell(ws1, 9, 3, 'קטגוריה מובילה')
        ws1.merge_cells('A10:B10')
        stat_val(ws1, 10, 1, f'₪{float(max_p["amount"]):,.0f} — {max_p["description"]}')
        ws1.merge_cells('C10:D10')
        stat_val(ws1, 10, 3, cats[0]['category'] if cats else '-')
        ws1.row_dimensions[10].height = 25

    # פילוח קטגוריות
    ws1.merge_cells('A12:D12')
    hdr_cell(ws1, 12, 1, 'פילוח לפי קטגוריה')
    ws1.row_dimensions[12].height = 22

    for i, cat in enumerate(cats):
        row = 13 + i
        pct = (float(cat['total']) / total * 100) if total > 0 else 0
        ws1.cell(row=row, column=1, value=cat['category']).font = Font(bold=True, size=10)
        ws1.cell(row=row, column=2, value=float(cat['total'])).number_format = '₪#,##0'
        ws1.cell(row=row, column=3, value=f'{pct:.1f}%').font = Font(color='6B7280', size=9)
        ws1.cell(row=row, column=1).alignment = Alignment(horizontal='right')

    # ── גרף עוגה — קטגוריות ── #
    if cats:
        pie = PieChart()
        pie.title = 'פילוח קטגוריות'
        pie.style = 10
        pie.width = 14
        pie.height = 10

        # נתוני קטגוריות לגרף בגיליון הסיכום
        data_ref = Reference(ws1, min_col=2, min_row=13, max_row=13 + len(cats) - 1)
        labels_ref = Reference(ws1, min_col=1, min_row=13, max_row=13 + len(cats) - 1)
        pie.add_data(data_ref)
        pie.set_categories(labels_ref)
        from openpyxl.chart.label import DataLabelList
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showCatName = True
        ws1.add_chart(pie, 'A' + str(14 + len(cats)))

    # ── גיליון 2: תשלומים ── #
    ws2 = wb.create_sheet('פירוט תשלומים')
    ws2.sheet_view.rightToLeft = True
    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 14
    ws2.column_dimensions['C'].width = 20
    ws2.column_dimensions['D'].width = 14

    for col, hdr in enumerate(['תיאור', 'סכום (₪)', 'קטגוריה', 'תאריך'], 1):
        hdr_cell(ws2, 1, col, hdr)
    ws2.row_dimensions[1].height = 22

    thin = Side(style='thin', color='E5E7EB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, p in enumerate(ps):
        row = 2 + i
        try:
            date_clean = '/'.join(reversed(p['date'].split(' ')[0].split('-')))
        except:
            date_clean = p['date']

        fill_color = 'F9FAFB' if i % 2 == 0 else 'FFFFFF'
        for col, val in enumerate([p['description'], float(p['amount']), p['category'], date_clean], 1):
            c = ws2.cell(row=row, column=col, value=val)
            c.fill = PatternFill('solid', fgColor=fill_color)
            c.border = border
            c.alignment = Alignment(horizontal='right' if col != 2 else 'center')
        ws2.cell(row=row, column=2).number_format = '₪#,##0.00'

    # שורת סיכום
    sum_row = 2 + len(ps)
    ws2.cell(row=sum_row, column=1, value='סה"כ').font = Font(bold=True)
    total_cell = ws2.cell(row=sum_row, column=2, value=total)
    total_cell.font = Font(bold=True, color='2563EB')
    total_cell.number_format = '₪#,##0.00'
    total_cell.fill = PatternFill('solid', fgColor='EFF6FF')

    # ── גיליון 3: הוצאות יומיות + גרף ── #
    ws3 = wb.create_sheet('גרף יומי')
    ws3.sheet_view.rightToLeft = True
    ws3.column_dimensions['A'].width = 10
    ws3.column_dimensions['B'].width = 16

    hdr_cell(ws3, 1, 1, 'יום')
    hdr_cell(ws3, 1, 2, 'סכום (₪)')
    ws3.row_dimensions[1].height = 22

    for day in range(1, days_in_month + 1):
        ws3.cell(row=day + 1, column=1, value=day)
        val = daily_map.get(day, 0)
        c = ws3.cell(row=day + 1, column=2, value=val)
        c.number_format = '₪#,##0'
        if val > 0:
            c.font = Font(color='2563EB', bold=True)

    bar = BarChart()
    bar.title = f'הוצאות יומיות — {month_name} {now.year}'
    bar.style = 10
    bar.type = 'col'
    bar.grouping = 'clustered'
    bar.width = 22
    bar.height = 14
    bar.shape = 4

    data = Reference(ws3, min_col=2, min_row=1, max_row=days_in_month + 1)
    cats_ref = Reference(ws3, min_col=1, min_row=2, max_row=days_in_month + 1)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats_ref)
    bar.series[0].graphicalProperties.solidFill = '2563EB'
    bar.series[0].graphicalProperties.line.solidFill = '1E3A5F'

    ws3.add_chart(bar, 'D1')

    m = io.BytesIO()
    wb.save(m)
    m.seek(0)
    return send_file(m, as_attachment=True,
                     download_name=f'payments_{cm}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/chart_data')
@require_auth
def chart_data():
    fid = get_family_id()
    now = now_israel()
    ms = get_cycle_month(fid)
    start, end, _ = get_cycle_range(fid)
    with get_db() as conn:
        cs = conn.execute(
            'SELECT p.category,SUM(p.amount) as total,c.color FROM payments p LEFT JOIN categories c ON p.category=c.name AND (c.family_id IS NULL OR c.family_id=p.family_id) WHERE p.month=? AND p.archived=FALSE AND p.family_id=? GROUP BY p.category ORDER BY total DESC',
            (ms, fid)).fetchall()
        dq = conn.execute('SELECT date,amount FROM payments WHERE month=? AND archived=FALSE AND family_id=?',
                          (ms, fid)).fetchall()
    # Aggregate by actual date
    sp = defaultdict(float)
    for r in dq:
        try:
            d = datetime.strptime(r['date'].split('.')[0], '%Y-%m-%d %H:%M:%S').date()
            sp[d] += float(r['amount'])
        except:
            pass
    # Generate all days in cycle range
    start_d = start.date() if isinstance(start, datetime) else start
    end_d = end.date() if isinstance(end, datetime) else end
    all_days = []
    cur = start_d
    while cur <= end_d:
        all_days.append(cur)
        cur += timedelta(days=1)
    today = now.date()
    today_index = next((i for i, d in enumerate(all_days) if d == today), -1)
    labels = [d.day for d in all_days]
    return jsonify({'categories': {'labels': [r['category'] for r in cs], 'data': [float(r['total']) for r in cs],
                                   'colors': [r['color'] or '#6c757d' for r in cs]},
                    'daily': {'labels': labels, 'data': [sp.get(d, 0) for d in all_days],
                              'today_index': today_index}})


@app.route('/history')
@require_auth
def history():
    fid = get_family_id()
    with get_db() as conn:
        cy = conn.execute(
            'SELECT label,total,count,archived_at FROM archived_cycles WHERE family_id=? ORDER BY archived_at DESC',
            (fid,)).fetchall()
        ap = conn.execute('SELECT * FROM payments WHERE archived=TRUE AND family_id=? ORDER BY date DESC LIMIT 100',
                          (fid,)).fetchall()
    return render_template('history.html', monthly_summaries=cy, archived_payments=ap)


@app.route('/api/history/data')
@require_auth
def history_data():
    fid = get_family_id()
    year = request.args.get('year', now_israel().year, type=int)
    cycle_day = get_cycle_day(fid)
    hebrew_months = {1:'ינואר',2:'פברואר',3:'מרץ',4:'אפריל',5:'מאי',6:'יוני',
                     7:'יולי',8:'אוגוסט',9:'ספטמבר',10:'אוקטובר',11:'נובמבר',12:'דצמבר'}
    reverse_hebrew = {'ינואר':1,'פברואר':2,'מרץ':3,'אפריל':4,'מאי':5,'יוני':6,
                       'יולי':7,'אוגוסט':8,'ספטמבר':9,'אוקטובר':10,'נובמבר':11,'דצמבר':12}
    with get_db() as conn:
        # Load archived labels for this family+year so past months keep original labels
        archived_labels = {}
        for arc in conn.execute(
                'SELECT id, month, label FROM archived_cycles WHERE family_id=?',
                (fid,)).fetchall():
            month_key = arc['month'] if arc['month'] else None
            # Backfill: parse month from label if month column is empty
            if not month_key and arc['label']:
                for heb_name, m_num in reverse_hebrew.items():
                    if arc['label'].startswith(heb_name):
                        parts = arc['label'].split()
                        for p in parts:
                            if p.isdigit() and len(p) == 4:
                                month_key = f'{p}-{m_num:02d}'
                                conn.execute('UPDATE archived_cycles SET month=? WHERE id=?', (month_key, arc['id']))
                                break
                        break
            if month_key and month_key.startswith(f'{year}-'):
                archived_labels[month_key] = arc['label']
        months = []
        for m in range(1, 13):
            month_str = f'{year}-{m:02d}'
            r = conn.execute(
                'SELECT COALESCE(SUM(amount),0) as total,COUNT(*) as count FROM payments WHERE month=? AND family_id=?',
                (month_str, fid)).fetchone()
            # Use stored label from archive if available, otherwise compute dynamically
            if month_str in archived_labels:
                label = archived_labels[month_str]
            elif cycle_day == 1:
                label = hebrew_months[m]
            else:
                nm = 1 if m == 12 else m + 1
                end_day = cycle_day - 1 if cycle_day > 1 else calendar.monthrange(year, nm)[1]
                label = f'{hebrew_months[m]} ({cycle_day}.{m}–{end_day}.{nm})'
            months.append({'month': m, 'total': float(r['total']), 'count': r['count'], 'label': label})
        ta = conn.execute('SELECT COALESCE(SUM(amount),0) as total FROM payments WHERE family_id=?', (fid,)).fetchone()[
            'total']
        mc = conn.execute('SELECT COUNT(DISTINCT month) as cnt FROM payments WHERE amount>0 AND family_id=?',
                          (fid,)).fetchone()['cnt']
    return jsonify(
        {'months': months, 'total_all': float(ta), 'avg_monthly': float(ta) / mc if mc else 0,
         'month_count': mc, 'cycle_day': cycle_day})


@app.route('/api/history/month')
@require_auth
def history_month_detail():
    fid = get_family_id()
    year = request.args.get('year', now_israel().year, type=int)
    month = request.args.get('month', now_israel().month, type=int)
    ms = f'{year}-{month:02d}'
    cycle_day = get_cycle_day(fid)
    hebrew_months = {1:'ינואר',2:'פברואר',3:'מרץ',4:'אפריל',5:'מאי',6:'יוני',
                     7:'יולי',8:'אוגוסט',9:'ספטמבר',10:'אוקטובר',11:'נובמבר',12:'דצמבר'}
    reverse_hebrew = {'ינואר':1,'פברואר':2,'מרץ':3,'אפריל':4,'מאי':5,'יוני':6,
                       'יולי':7,'אוגוסט':8,'ספטמבר':9,'אוקטובר':10,'נובמבר':11,'דצמבר':12}
    with get_db() as conn:
        # Check if this month has a stored archived label (by month column or by parsing label)
        arc = conn.execute('SELECT id, month, label FROM archived_cycles WHERE family_id=? AND month=?',
                           (fid, ms)).fetchone()
        if not arc:
            # Fallback: search by label text for old records without month column
            all_arcs = conn.execute('SELECT id, month, label FROM archived_cycles WHERE family_id=? AND (month IS NULL OR month="")',
                                    (fid,)).fetchall()
            for a in all_arcs:
                for heb_name, m_num in reverse_hebrew.items():
                    if a['label'] and a['label'].startswith(heb_name):
                        parts = a['label'].split()
                        for p in parts:
                            if p.isdigit() and len(p) == 4:
                                parsed_ms = f'{p}-{m_num:02d}'
                                conn.execute('UPDATE archived_cycles SET month=? WHERE id=?', (parsed_ms, a['id']))
                                if parsed_ms == ms:
                                    arc = a
                                break
                        break
    if arc and arc['label']:
        cycle_label = arc['label']
        # Extract dates from label if present (format: "מרץ 2026 (10.3–9.4)")
        dates_match = re.search(r'\((.+)\)', cycle_label)
        cycle_dates = dates_match.group(1) if dates_match else ''
    elif cycle_day == 1:
        cycle_label = f'{hebrew_months[month]} {year}'
        last_day = calendar.monthrange(year, month)[1]
        cycle_dates = f'1.{month} – {last_day}.{month}'
    else:
        nm = 1 if month == 12 else month + 1
        end_day = cycle_day - 1
        cycle_label = f'{hebrew_months[month]} {year} ({cycle_day}.{month}–{end_day}.{nm})'
        cycle_dates = f'{cycle_day}.{month} – {end_day}.{nm}'
    with get_db() as conn:
        ps = conn.execute(
            'SELECT p.id,p.description,p.amount,p.category,p.date,COALESCE(c.color,\'#6c757d\') as color FROM payments p LEFT JOIN categories c ON p.category=c.name AND (c.family_id IS NULL OR c.family_id=p.family_id) WHERE p.month=? AND p.family_id=? ORDER BY p.date DESC',
            (ms, fid)).fetchall()
        t = conn.execute(
            'SELECT COALESCE(SUM(amount),0) as total,COUNT(*) as count FROM payments WHERE month=? AND family_id=?',
            (ms, fid)).fetchone()
        ca = conn.execute(
            'SELECT p.category,SUM(p.amount) as total,COALESCE(c.color,\'#6c757d\') as color FROM payments p LEFT JOIN categories c ON p.category=c.name AND (c.family_id IS NULL OR c.family_id=p.family_id) WHERE p.month=? AND p.family_id=? GROUP BY p.category ORDER BY total DESC',
            (ms, fid)).fetchall()
    return jsonify({'payments': [
        {'id': p['id'], 'description': p['description'], 'amount': float(p['amount']), 'category': p['category'],
         'color': p['color'], 'date': p['date'].split(' ')[0] if p['date'] else ''} for p in ps],
        'categories': [{'name': c['category'], 'total': float(c['total']), 'color': c['color']} for c in
                       ca], 'total': float(t['total']), 'count': t['count'],
        'cycle_label': cycle_label, 'cycle_dates': cycle_dates})


@app.route('/shopping-list')
@require_auth
def shopping_list(): return render_template('shopping_list.html')


@app.route('/api/shopping-items', methods=['GET'])
@require_auth
def get_shopping_items():
    fid = get_family_id()
    with get_db() as conn:
        items = conn.execute(
            'SELECT id,name,quantity,checked,image,COALESCE(favorite,0) as favorite,COALESCE(category,"") as category FROM shopping_items WHERE family_id=? ORDER BY checked ASC, category ASC, created_at DESC',
            (fid,)).fetchall()
    return jsonify([dict(i) for i in items])


@app.route('/api/shopping-items', methods=['POST'])
@require_auth
def add_shopping_item():
    fid = get_family_id()
    data = request.get_json()
    name = data.get('name', '').strip()
    if not name: return jsonify({'error': 'Name required'}), 400
    cat = data.get('category', '')
    with get_db() as conn:
        cur = conn.execute(
            'INSERT INTO shopping_items (family_id,name,quantity,checked,category) VALUES (?,?,?,FALSE,?)',
            (fid, name, data.get('quantity', 1), cat))

    # Push notification to family
    user_id = request.api_user['user_id'] if hasattr(request, 'api_user') else session.get('user_id')
    user_name = request.api_user['display_name'] if hasattr(request, 'api_user') else session.get('display_name', '')
    send_push_to_family(fid,
        '🛒 פריט חדש ברשימה',
        f'{user_name} הוסיף/ה: {name}',
        exclude_user_id=user_id)

    return jsonify({'id': cur.lastrowid}), 201


@app.route('/api/shopping-items/<int:iid>', methods=['PUT'])
@require_auth
def update_shopping_item(iid):
    fid = get_family_id()
    data = request.get_json()
    allowed = ['checked', 'name', 'quantity', 'image', 'favorite', 'category']
    with get_db() as conn:
        for f in allowed:
            if f in data: conn.execute(f'UPDATE shopping_items SET {f}=? WHERE id=? AND family_id=?',
                                       (data[f], iid, fid))
        if data.get('favorite'):
            item = conn.execute(
                'SELECT name,quantity,COALESCE(category,"") as category FROM shopping_items WHERE id=? AND family_id=?',
                (iid, fid)).fetchone()
            if item: conn.execute(
                'INSERT OR REPLACE INTO shopping_favorites (family_id,name,quantity,category) VALUES (?,?,?,?)',
                (fid, item['name'], item['quantity'], item['category']))
        elif 'favorite' in data and not data['favorite']:
            item = conn.execute('SELECT name FROM shopping_items WHERE id=? AND family_id=?', (iid, fid)).fetchone()
            if item: conn.execute('DELETE FROM shopping_favorites WHERE family_id=? AND name=?', (fid, item['name']))

    return jsonify({'success': True})


@app.route('/api/shopping-items/favorites', methods=['GET'])
@require_auth
def get_favorites():
    fid = get_family_id()
    with get_db() as conn:
        favs = conn.execute(
            'SELECT name,quantity,category FROM shopping_favorites WHERE family_id=? ORDER BY category,name',
            (fid,)).fetchall()
    return jsonify([dict(f) for f in favs])


@app.route('/api/shopping-items/add-favorites', methods=['POST'])
@require_auth
def add_favorites():
    fid = get_family_id()
    with get_db() as conn:
        favs = conn.execute('SELECT name,quantity,category FROM shopping_favorites WHERE family_id=?',
                            (fid,)).fetchall()
        added = 0
        names = []
        for f in favs:
            existing = conn.execute('SELECT id FROM shopping_items WHERE family_id=? AND name=? AND checked=FALSE',
                                    (fid, f['name'])).fetchone()
            if not existing:
                conn.execute(
                    'INSERT INTO shopping_items (family_id,name,quantity,checked,favorite,category) VALUES (?,?,?,FALSE,TRUE,?)',
                    (fid, f['name'], f['quantity'], f['category']))
                added += 1
                names.append(f['name'])

    if added > 0:
        user_id = request.api_user['user_id'] if hasattr(request, 'api_user') else session.get('user_id')
        user_name = request.api_user['display_name'] if hasattr(request, 'api_user') else session.get('display_name', '')
        items_text = ', '.join(names[:3])
        if len(names) > 3:
            items_text += f' ועוד {len(names)-3}'
        send_push_to_family(fid, '🛒 מועדפים נוספו לרשימה',
            f'{user_name} הוסיף/ה {added} פריטים: {items_text}',
            exclude_user_id=user_id)

    return jsonify({'success': True, 'added': added})


@app.route('/api/shopping-items/delete-favorite', methods=['POST'])
@require_auth
def delete_favorite():
    fid = get_family_id()
    data = request.get_json()
    name = data.get('name', '')
    with get_db() as conn:
        conn.execute('DELETE FROM shopping_favorites WHERE family_id=? AND name=?', (fid, name))
        conn.execute('UPDATE shopping_items SET favorite=FALSE WHERE family_id=? AND name=?', (fid, name))
    return jsonify({'success': True})


@app.route('/api/shopping-items/add-new-favorite', methods=['POST'])
@require_auth
def add_new_favorite():
    fid = get_family_id()
    data = request.get_json()
    name = data.get('name', '').strip()
    qty = data.get('quantity', 1)
    cat = data.get('category', '')
    if not name: return jsonify({'error': 'Missing name'}), 400
    with get_db() as conn:
        conn.execute('INSERT OR REPLACE INTO shopping_favorites (family_id,name,quantity,category) VALUES (?,?,?,?)',
                     (fid, name, qty, cat))
    return jsonify({'success': True}), 201


@app.route('/api/shopping-items/edit-favorite', methods=['POST'])
@require_auth
def edit_favorite():
    fid = get_family_id()
    data = request.get_json()
    old_name = data.get('old_name', '')
    name = data.get('name', '').strip()
    qty = data.get('quantity', 1)
    cat = data.get('category', '')
    if not name: return jsonify({'error': 'Missing name'}), 400
    with get_db() as conn:
        conn.execute('DELETE FROM shopping_favorites WHERE family_id=? AND name=?', (fid, old_name))
        conn.execute('INSERT OR REPLACE INTO shopping_favorites (family_id,name,quantity,category) VALUES (?,?,?,?)',
                     (fid, name, qty, cat))
    return jsonify({'success': True})


@app.route('/api/shopping-items/<int:iid>', methods=['DELETE'])
@require_auth
def delete_shopping_item(iid):
    fid = get_family_id()
    with get_db() as conn: conn.execute('DELETE FROM shopping_items WHERE id=? AND family_id=?', (iid, fid))
    return jsonify({'success': True})


@app.route('/api/shopping-items/clear-completed', methods=['DELETE'])
@require_auth
def clear_completed_items():
    fid = get_family_id()
    with get_db() as conn:
        count = conn.execute('SELECT COUNT(*) as c FROM shopping_items WHERE family_id=? AND checked=TRUE', (fid,)).fetchone()['c']
        conn.execute('DELETE FROM shopping_items WHERE family_id=? AND checked=TRUE', (fid,))

    if count > 0:
        user_id = request.api_user['user_id'] if hasattr(request, 'api_user') else session.get('user_id')
        user_name = request.api_user['display_name'] if hasattr(request, 'api_user') else session.get('display_name', '')
        send_push_to_family(fid, '🧹 רשימה נוקתה',
            f'{user_name} ניקה/תה {count} פריטים שהושלמו',
            exclude_user_id=user_id)

    return jsonify({'success': True})


@app.route('/api/recurring', methods=['GET'])
@require_auth
def get_recurring():
    fid = get_family_id()
    with get_db() as conn: items = conn.execute(
        'SELECT id,description,amount,category FROM recurring_payments WHERE family_id=? ORDER BY category,description',
        (fid,)).fetchall()
    return jsonify([dict(i) for i in items])


@app.route('/api/recurring', methods=['POST'])
@require_auth
def add_recurring():
    fid = get_family_id()
    data = request.get_json()
    with get_db() as conn: conn.execute(
        'INSERT INTO recurring_payments (family_id,description,amount,category) VALUES (?,?,?,?)',
        (fid, data['description'], data['amount'], data.get('category', 'כללי')))
    return jsonify({'success': True}), 201


@app.route('/api/recurring/<int:rid>', methods=['DELETE'])
@require_auth
def delete_recurring(rid):
    fid = get_family_id()
    with get_db() as conn: conn.execute('DELETE FROM recurring_payments WHERE id=? AND family_id=?', (rid, fid))
    return jsonify({'success': True})

@app.route('/api/recurring/<int:rid>', methods=['PUT'])
@require_auth
def update_recurring(rid):
    fid=get_family_id(); data=request.get_json()
    with get_db() as conn:
        conn.execute('UPDATE recurring_payments SET description=?,amount=?,category=? WHERE id=? AND family_id=?',
                     (data['description'],data['amount'],data.get('category','כללי'),rid,fid))
    return jsonify({'success':True})

@app.route('/api/recurring/<int:rid>/add', methods=['POST'])
@require_auth
def add_recurring_to_month(rid):
    fid = get_family_id()
    cm = get_cycle_month(fid)
    cy = int(cm.split('-')[0])
    with get_db() as conn:
        r = conn.execute('SELECT * FROM recurring_payments WHERE id=? AND family_id=?', (rid, fid)).fetchone()
        if r: conn.execute(
            'INSERT INTO payments (family_id,description,amount,category,month,year,date) VALUES (?,?,?,?,?,?,?)',
            (fid, r['description'], r['amount'], r['category'], cm, cy,
             now_israel().strftime('%Y-%m-%d %H:%M:%S')))

    # Push + budget check
    if r:
        user_id = request.api_user['user_id'] if hasattr(request, 'api_user') else session.get('user_id')
        user_name = request.api_user['display_name'] if hasattr(request, 'api_user') else session.get('display_name', '')
        send_push_to_family(fid, '💰 תשלום קבוע נוסף',
            f'{user_name} הוסיף/ה: {r["description"]} — ₪{float(r["amount"]):.0f}',
            exclude_user_id=user_id)
        check_budget_alerts(fid)

    return jsonify({'success': True})


@app.route('/api/recurring/add-all', methods=['POST'])
@require_auth
def add_all_recurring():
    fid = get_family_id()
    cm = get_cycle_month(fid)
    cy = int(cm.split('-')[0])
    with get_db() as conn:
        items = conn.execute('SELECT * FROM recurring_payments WHERE family_id=?', (fid,)).fetchall()
        total_amount = 0
        for r in items:
            conn.execute(
                'INSERT INTO payments (family_id,description,amount,category,month,year,date) VALUES (?,?,?,?,?,?,?)',
                (fid, r['description'], r['amount'], r['category'], cm, cy,
                 now_israel().strftime('%Y-%m-%d %H:%M:%S')))
            total_amount += r['amount']

    if items:
        user_id = request.api_user['user_id'] if hasattr(request, 'api_user') else session.get('user_id')
        user_name = request.api_user['display_name'] if hasattr(request, 'api_user') else session.get('display_name', '')
        send_push_to_family(fid, '💰 תשלומים קבועים נוספו',
            f'{user_name} הוסיף/ה {len(items)} תשלומים קבועים — סה"כ ₪{total_amount:.0f}',
            exclude_user_id=user_id)
        check_budget_alerts(fid)

    return jsonify({'success': True, 'count': len(items)})


@app.route('/baby-tracker')
@require_auth
def baby_tracker(): return render_template('baby_tracker.html')


@app.route('/api/feedings', methods=['POST'])
@require_auth
def add_feeding():
    fid = get_family_id()
    data = request.get_json()
    ft = data.get('feeding_type', '')
    ct = data.get('custom_time', '')
    ds = f'{now_israel().strftime("%Y-%m-%d")} {ct}:00' if ct else now_israel().strftime('%Y-%m-%d %H:%M:%S')
    with get_db() as conn:
        cur = conn.execute(
            'INSERT INTO feedings (family_id,feeding_type,amount,duration,notes,date) VALUES (?,?,?,?,?,?)',
            (fid, ft, data.get('amount', 0), data.get('duration', 0), data.get('notes', ''), ds))

    # Push notification to family
    feeding_names = {
        'bottle': '🍼 בקבוק', 'breastfeeding': '🤱 הנקה', 'solid': '🥣 מוצק',
        'diaper': '🚼 חיתול', 'sleep': '😴 שינה', 'medication': '💊 תרופה'
    }
    feed_label = feeding_names.get(ft, ft)
    amount_val = data.get('amount', 0)
    detail = f' — {int(amount_val)} מ"ל' if ft == 'bottle' and amount_val else ''

    user_id = request.api_user['user_id'] if hasattr(request, 'api_user') else session.get('user_id')
    user_name = request.api_user['display_name'] if hasattr(request, 'api_user') else session.get('display_name', '')
    send_push_to_family(fid,
        f'👶 {feed_label}',
        f'{user_name} הוסיף/ה {feed_label}{detail}',
        exclude_user_id=user_id)

    return jsonify({'success': True, 'id': cur.lastrowid}), 201


@app.route('/api/feedings/<int:feed_id>', methods=['DELETE'])
@require_auth
def delete_feeding(feed_id):
    fid = get_family_id()
    with get_db() as conn: conn.execute('DELETE FROM feedings WHERE id=? AND family_id=?', (feed_id, fid))
    return jsonify({'success': True})


@app.route('/api/feedings/<int:feed_id>', methods=['PUT'])
@require_auth
def update_feeding(feed_id):
    fid = get_family_id()
    data = request.get_json()
    with get_db() as conn:
        if 'amount' in data:
            conn.execute('UPDATE feedings SET amount=? WHERE id=? AND family_id=?', (data['amount'], feed_id, fid))
        if 'notes' in data:
            conn.execute('UPDATE feedings SET notes=? WHERE id=? AND family_id=?', (data['notes'], feed_id, fid))
        if 'time' in data and data['time']:
            # Update time portion of date
            existing = conn.execute('SELECT date FROM feedings WHERE id=? AND family_id=?', (feed_id, fid)).fetchone()
            if existing:
                date_part = existing['date'].split(' ')[0]
                new_date = f'{date_part} {data["time"]}:00'
                conn.execute('UPDATE feedings SET date=? WHERE id=? AND family_id=?', (new_date, feed_id, fid))
    return jsonify({'success': True})


@app.route('/api/feedings/data')
@require_auth
def feedings_data():
    fid = get_family_id()
    qd = request.args.get('date', now_israel().strftime('%Y-%m-%d'))
    wo = int(request.args.get('week_offset', 0))
    today = now_israel().strftime('%Y-%m-%d')
    with get_db() as conn:
        df = conn.execute(
            'SELECT id,feeding_type,amount,duration,notes,date FROM feedings WHERE family_id=? AND date(date)=? ORDER BY date DESC',
            (fid, qd)).fetchall()
        # Rich stats for today
        bottles = conn.execute(
            'SELECT COUNT(*) as c,COALESCE(SUM(amount),0) as ml FROM feedings WHERE family_id=? AND date(date)=? AND feeding_type=?',
            (fid, qd, 'bottle')).fetchone()
        bf = conn.execute('SELECT COUNT(*) as c FROM feedings WHERE family_id=? AND date(date)=? AND feeding_type=?',
                          (fid, qd, 'breastfeeding')).fetchone()
        solids = conn.execute(
            'SELECT COUNT(*) as c FROM feedings WHERE family_id=? AND date(date)=? AND feeding_type=?',
            (fid, qd, 'solid')).fetchone()
        diapers = conn.execute(
            'SELECT COUNT(*) as c FROM feedings WHERE family_id=? AND date(date)=? AND feeding_type=?',
            (fid, qd, 'diaper')).fetchone()
        meds = conn.execute('SELECT COUNT(*) as c FROM feedings WHERE family_id=? AND date(date)=? AND feeding_type=?',
                            (fid, qd, 'medication')).fetchone()
        sleeps = conn.execute(
            'SELECT COUNT(*) as c,COALESCE(SUM(amount),0) as mins FROM feedings WHERE family_id=? AND date(date)=? AND feeding_type=?',
            (fid, qd, 'sleep')).fetchone()
        lf = conn.execute(
            'SELECT date FROM feedings WHERE family_id=? AND date(date)=? AND feeding_type IN (?,?,?) ORDER BY time(date) DESC LIMIT 1',
            (fid, qd, 'bottle', 'breastfeeding', 'solid')).fetchone()
        # Weekly
        cur = now_israel()
        dss = (cur.weekday() + 1) % 7
        ws = cur - timedelta(days=dss) + timedelta(weeks=wo)
        weekly = []
        dns = ['ראשון', 'שני', 'שלישי', 'רביעי', 'חמישי', 'שישי', 'שבת']
        chart_type = request.args.get('chart_type', 'bottle')
        for i in range(7):
            dd = ws + timedelta(days=i)
            d = dd.strftime('%Y-%m-%d')
            if chart_type == 'all':
                dr = conn.execute(
                    'SELECT COUNT(*) as count, 0 as total_amount FROM feedings WHERE family_id=? AND date(date)=?',
                    (fid, d)).fetchone()
            else:
                dr = conn.execute(
                    'SELECT COUNT(*) as count, COALESCE(SUM(amount),0) as total_amount FROM feedings WHERE family_id=? AND date(date)=? AND feeding_type=?',
                    (fid, d, chart_type)).fetchone()
            dn = dns[i]
            if d == today: dn = 'היום'
            weekly.append({'date': d, 'label': dn, 'short': dd.strftime('%d/%m'), 'count': dr['count'],
                           'total_amount': float(dr['total_amount'])})
    # Format
    fmt = []
    for f in df:
        try:
            ts = datetime.strptime(f['date'].split('.')[0], '%Y-%m-%d %H:%M:%S').strftime('%H:%M')
        except:
            ts = ''
        fmt.append(
            {'id': f['id'], 'feeding_type': f['feeding_type'], 'amount': float(f['amount']), 'duration': f['duration'],
             'notes': f['notes'], 'time': ts})
    lt = '--'
    if lf:
        try:
            ldt = datetime.strptime(lf['date'].split('.')[0], '%Y-%m-%d %H:%M:%S')
            mins = int((now_israel() - ldt).total_seconds() // 60)
            if mins < 0:
                lt = 'עוד מעט'
            elif mins < 60:
                lt = f'{mins} דק\''
            else:
                lt = f'{mins // 60}:{mins % 60:02d}h'
        except:
            pass
    return jsonify({'today_feedings': fmt, 'query_date': qd, 'is_today': qd == today,
                    'stats': {'bottles': bottles['c'], 'bottle_ml': float(bottles['ml']), 'breastfeedings': bf['c'],
                              'solids': solids['c'], 'diapers': diapers['c'], 'medications': meds['c'],
                              'sleeps': sleeps['c'], 'sleep_mins': float(sleeps['mins']), 'last_feeding': lt},
                    'weekly': weekly})


# ══════════════════════════════════════════════
# MOBILE API ENDPOINTS (JWT Auth)
# ══════════════════════════════════════════════

# --- AUTH: LOGIN ---
@app.route('/api/auth/login', methods=['POST'])
@csrf.exempt
def api_login():
    data = request.get_json(silent=True) or {}
    email = data.get('email', '').strip()
    password = data.get('password', '')
    # Backward compat: support username login too
    username = data.get('username', '').strip().lower()

    if not password or (not email and not username):
        return jsonify({'error': 'אימייל/שם משתמש וסיסמה נדרשים'}), 400

    # If username provided, look up email first
    if not email and username:
        with get_db() as conn:
            u = conn.execute('SELECT email, password_hash FROM users WHERE username=?', (username,)).fetchone()
            if u and u['email']:
                email = u['email']
            elif u and check_password_hash(u['password_hash'], password):
                # Fallback: local auth for users without email in Firebase
                user = conn.execute('SELECT * FROM users WHERE username=?', (username,)).fetchone()
                token = create_jwt_token(user['id'], user['username'],
                                          user['display_name'] or user['username'],
                                          user['family_id'], bool(user['is_admin']))
                return jsonify({
                    'token': token,
                    'user': {'id': user['id'], 'username': user['username'],
                             'display_name': user['display_name'] or user['username'],
                             'email': user['email'], 'family_id': user['family_id'],
                             'is_admin': bool(user['is_admin'])}
                })
            else:
                return jsonify({'error': 'שם משתמש או סיסמה שגויים'}), 401

    # Firebase Auth verification
    fb_uid, error = firebase_verify_login(email, password)
    if error:
        return jsonify({'error': error}), 401

    with get_db() as conn:
        user = conn.execute('SELECT * FROM users WHERE email=? OR firebase_uid=?',
                            (email, fb_uid or '')).fetchone()
        if not user:
            return jsonify({'error': 'משתמש לא נמצא. יש להירשם קודם'}), 401
        if fb_uid and not user['firebase_uid']:
            conn.execute('UPDATE users SET firebase_uid=? WHERE id=?', (fb_uid, user['id']))
        token = create_jwt_token(user['id'], user['username'],
                                  user['display_name'] or user['username'],
                                  user['family_id'], bool(user['is_admin']))
        return jsonify({
            'token': token,
            'user': {'id': user['id'], 'username': user['username'],
                     'display_name': user['display_name'] or user['username'],
                     'email': user['email'], 'family_id': user['family_id'],
                     'is_admin': bool(user['is_admin'])}
        })


# --- AUTH: REGISTER ---
@app.route('/api/auth/register', methods=['POST'])
@csrf.exempt
def api_register():
    data = request.get_json(silent=True) or {}
    display_name = data.get('display_name', '').strip()
    username = data.get('username', '').strip().lower()
    email = data.get('email', '').strip()
    password = data.get('password', '')
    password2 = data.get('password2', '')
    if not display_name:
        return jsonify({'error': 'שם תצוגה נדרש'}), 400
    if not re.match(r'^[a-zA-Z0-9_]{3,20}$', username):
        return jsonify({'error': 'שם משתמש חייב להכיל 3-20 תווים באנגלית, מספרים או _'}), 400
    if email and not re.match(r'^[^@]+@[^@]+\.[^@]+$', email):
        return jsonify({'error': 'כתובת אימייל לא תקינה'}), 400
    if len(password) < 6:
        return jsonify({'error': 'סיסמה חייבת להיות לפחות 6 תווים'}), 400
    if password != password2:
        return jsonify({'error': 'הסיסמאות לא תואמות'}), 400
    with get_db() as conn:
        if conn.execute('SELECT id FROM users WHERE username=?', (username,)).fetchone():
            return jsonify({'error': 'שם משתמש כבר תפוס'}), 409
        if email and conn.execute('SELECT id FROM users WHERE email=?', (email,)).fetchone():
            return jsonify({'error': 'כתובת אימייל כבר רשומה'}), 409
        # Create in Firebase Auth
        fb_uid, fb_error = firebase_create_user(email, password, display_name)
        if fb_error:
            return jsonify({'error': fb_error}), 409
        conn.execute('INSERT INTO users (username,email,display_name,password_hash,firebase_uid) VALUES (?,?,?,?,?)',
                     (username, email, display_name, generate_password_hash(password), fb_uid or ''))
        user = conn.execute('SELECT * FROM users WHERE username=?', (username,)).fetchone()
        token = create_jwt_token(user['id'], username, display_name, None, False)
        return jsonify({
            'token': token,
            'user': {'id': user['id'], 'username': username, 'display_name': display_name,
                     'email': email, 'family_id': None, 'is_admin': False}
        }), 201


# --- AUTH: GET CURRENT USER ---
@app.route('/api/auth/me', methods=['GET'])
@csrf.exempt
@require_api_auth
def api_me():
    user = request.api_user
    family = None
    members = []
    if user['family_id']:
        with get_db() as conn:
            fam = conn.execute('SELECT * FROM families WHERE id=?', (user['family_id'],)).fetchone()
            if fam:
                family = {'id': fam['id'], 'name': fam['name'], 'invite_code': fam['invite_code']}
            mems = conn.execute('SELECT id,display_name,username FROM users WHERE family_id=?',
                                (user['family_id'],)).fetchall()
            members = [{'id': m['id'], 'display_name': m['display_name'], 'username': m['username']} for m in mems]
    return jsonify({'user': user, 'family': family, 'members': members})


# --- AUTH: REFRESH TOKEN ---
@app.route('/api/auth/refresh', methods=['POST'])
@csrf.exempt
@require_api_auth
def api_refresh_token():
    u = request.api_user
    token = create_jwt_token(u['user_id'], u['username'], u['display_name'], u['family_id'], u['is_admin'])
    return jsonify({'token': token})


# --- AUTH: CHANGE PASSWORD ---
@app.route('/api/auth/change-password', methods=['POST'])
@csrf.exempt
@require_api_auth
def api_change_password():
    data = request.get_json(silent=True) or {}
    current_pw = data.get('current_password', '')
    new_pw = data.get('new_password', '')
    if len(new_pw) < 6:
        return jsonify({'error': 'סיסמה חדשה חייבת להיות לפחות 6 תווים'}), 400
    with get_db() as conn:
        user = conn.execute('SELECT * FROM users WHERE id=?', (request.api_user['user_id'],)).fetchone()
        if not user or not check_password_hash(user['password_hash'], current_pw):
            return jsonify({'error': 'סיסמה נוכחית שגויה'}), 401
        conn.execute('UPDATE users SET password_hash=? WHERE id=?',
                     (generate_password_hash(new_pw), request.api_user['user_id']))
        # Sync to Firebase Auth
        firebase_update_password(user['email'] or user['username'], new_pw)
        # שליחת מייל התראה על שינוי סיסמה
        try:
            if user['email']:
                msg = Message(
                    subject='סיסמתך שונתה — OurHome IL',
                    recipients=[user['email']],
                    html=f'''
                    <div dir="rtl" style="font-family:Arial;max-width:500px;margin:0 auto;">
                        <h2>סיסמה שונתה</h2>
                        <p>הסיסמה לחשבון שלך באפליקציית OurHome IL שונתה זה עתה.</p>
                        <p style="color:#888;font-size:0.85rem;">
                            אם לא ביקשת שינוי זה — פנה אלינו מיד.
                        </p>
                    </div>
                    '''
                )
                mail.send(msg)
        except Exception as e:
            print(f'Mail error: {e}')
    return jsonify({'message': 'סיסמה שונתה בהצלחה'})


# --- AUTH: FORGOT PASSWORD ---
@app.route('/api/auth/forgot-password', methods=['POST'])
@csrf.exempt
def api_forgot_password():
    data = request.get_json(silent=True) or {}
    email = data.get('email', '').strip()
    if not email:
        return jsonify({'error': 'אימייל נדרש'}), 400
    firebase_send_reset_email(email)
    return jsonify({'message': 'אם האימייל קיים במערכת, נשלחה הודעה עם קישור לאיפוס'})


# --- AUTH: RESET PASSWORD ---
@app.route('/api/auth/reset-password', methods=['POST'])
@csrf.exempt
def api_reset_password():
    data = request.get_json(silent=True) or {}
    token = data.get('token', '')
    password = data.get('password', '')
    password2 = data.get('password2', '')
    if not token:
        return jsonify({'error': 'טוקן נדרש'}), 400
    if len(password) < 6:
        return jsonify({'error': 'סיסמה חייבת להיות לפחות 6 תווים'}), 400
    if password != password2:
        return jsonify({'error': 'הסיסמאות לא תואמות'}), 400
    with get_db() as conn:
        user = conn.execute('SELECT * FROM users WHERE reset_token=? AND reset_token_exp > ?',
                            (token, now_israel().strftime('%Y-%m-%d %H:%M:%S'))).fetchone()
        if not user:
            return jsonify({'error': 'הקישור לא תקין או פג תוקף'}), 400
        conn.execute('UPDATE users SET password_hash=?, reset_token="", reset_token_exp=NULL WHERE id=?',
                     (generate_password_hash(password), user['id']))
    # Sync to Firebase Auth
    firebase_update_password(user['email'] or user['username'], password)
    return jsonify({'message': 'סיסמה שונתה בהצלחה'})


# --- FAMILY: GET INFO ---
@app.route('/api/family/info', methods=['GET'])
@csrf.exempt
@require_api_auth
def api_family_info():
    user = request.api_user
    if not user['family_id']:
        return jsonify({'family': None, 'members': []})
    with get_db() as conn:
        family = conn.execute('SELECT * FROM families WHERE id=?', (user['family_id'],)).fetchone()
        members = conn.execute('SELECT id,display_name,username FROM users WHERE family_id=?',
                               (user['family_id'],)).fetchall()
    return jsonify({
        'family': {'id': family['id'], 'name': family['name'], 'invite_code': family['invite_code']} if family else None,
        'members': [{'id': m['id'], 'display_name': m['display_name'], 'username': m['username']} for m in members]
    })


# --- FAMILY: CREATE ---
@app.route('/api/family/create', methods=['POST'])
@csrf.exempt
@require_api_auth
def api_create_family():
    data = request.get_json(silent=True) or {}
    name = data.get('family_name', '').strip()
    if not name:
        return jsonify({'error': 'שם משפחה נדרש'}), 400
    code = generate_invite_code()
    uid = request.api_user['user_id']
    with get_db() as conn:
        conn.execute('INSERT INTO families (name,invite_code,created_by) VALUES (?,?,?)', (name, code, uid))
        fam = conn.execute('SELECT * FROM families WHERE invite_code=?', (code,)).fetchone()
        conn.execute('UPDATE users SET family_id=? WHERE id=?', (fam['id'], uid))
    u = request.api_user
    token = create_jwt_token(uid, u['username'], u['display_name'], fam['id'], u['is_admin'])
    return jsonify({
        'token': token,
        'family': {'id': fam['id'], 'name': name, 'invite_code': code},
        'message': f'משפחה "{name}" נוצרה!'
    }), 201


# --- FAMILY: JOIN ---
@app.route('/api/family/join', methods=['POST'])
@csrf.exempt
@require_api_auth
def api_join_family():
    data = request.get_json(silent=True) or {}
    code = data.get('invite_code', '').strip().upper()
    if not code:
        return jsonify({'error': 'קוד הזמנה נדרש'}), 400
    uid = request.api_user['user_id']
    with get_db() as conn:
        fam = conn.execute('SELECT * FROM families WHERE invite_code=?', (code,)).fetchone()
        if not fam:
            return jsonify({'error': 'קוד הזמנה לא נמצא'}), 404
        conn.execute('UPDATE users SET family_id=? WHERE id=?', (fam['id'], uid))
    u = request.api_user
    token = create_jwt_token(uid, u['username'], u['display_name'], fam['id'], u['is_admin'])

    # Notify family about new member
    send_push_to_family(fam['id'], '👨‍👩‍👧 חבר חדש במשפחה!',
        f'{u["display_name"]} הצטרף/ה למשפחה',
        exclude_user_id=uid)

    return jsonify({
        'token': token,
        'family': {'id': fam['id'], 'name': fam['name'], 'invite_code': fam['invite_code']},
        'message': f'הצטרפת למשפחת {fam["name"]}!'
    })

# --- SETTINGS: GET ---
@app.route('/api/settings', methods=['GET'])
@csrf.exempt
@require_api_auth
def api_get_settings():
    user = request.api_user
    family = None
    members = []
    if user['family_id']:
        with get_db() as conn:
            fam = conn.execute('SELECT * FROM families WHERE id=?', (user['family_id'],)).fetchone()
            if fam:
                family = {'id': fam['id'], 'name': fam['name'], 'invite_code': fam['invite_code']}
            mems = conn.execute('SELECT id,display_name,username FROM users WHERE family_id=?',
                                (user['family_id'],)).fetchall()
            members = [{'id': m['id'], 'display_name': m['display_name'], 'username': m['username']} for m in mems]
    return jsonify({
        'user': {'id': user['user_id'], 'username': user['username'],
                 'display_name': user['display_name'], 'email': user['email']},
        'family': family, 'members': members
    })


# --- SETTINGS: UPDATE PROFILE ---
@app.route('/api/settings/profile', methods=['PUT'])
@csrf.exempt
@require_api_auth
def api_update_profile():
    data = request.get_json(silent=True) or {}
    display_name = data.get('display_name', '').strip()
    if not display_name:
        return jsonify({'error': 'שם תצוגה נדרש'}), 400
    uid = request.api_user['user_id']
    with get_db() as conn:
        conn.execute('UPDATE users SET display_name=? WHERE id=?', (display_name, uid))
    u = request.api_user
    token = create_jwt_token(uid, u['username'], display_name, u['family_id'], u['is_admin'])
    return jsonify({'token': token, 'message': 'שם תצוגה עודכן!', 'display_name': display_name})


# --- PAYMENTS: DELETE ---
@app.route('/api/payments/<int:pid>', methods=['DELETE'])
@csrf.exempt
@require_api_family
def api_delete_payment(pid):
    fid = request.api_user['family_id']
    with get_db() as conn:
        p = conn.execute('SELECT * FROM payments WHERE id=? AND family_id=?', (pid, fid)).fetchone()
        if not p:
            return jsonify({'error': 'תשלום לא נמצא'}), 404
        conn.execute('DELETE FROM payments WHERE id=? AND family_id=?', (pid, fid))

    user_id = request.api_user['user_id']
    user_name = request.api_user['display_name']
    send_push_to_family(fid,
        '🗑️ תשלום נמחק',
        f'{user_name} מחק/ה: {p["description"]} — ₪{float(p["amount"]):.0f}',
        exclude_user_id=user_id)
    return jsonify({'message': 'תשלום נמחק'})


# --- PAYMENTS: ARCHIVE MONTH ---
@app.route('/api/payments/archive', methods=['POST'])
@csrf.exempt
@require_api_family
def api_archive_month():
    fid = request.api_user['family_id']
    now = now_israel()
    cm = get_cycle_month(fid)
    start, end, label = get_cycle_range(fid)
    with get_db() as conn:
        payments = conn.execute('SELECT * FROM payments WHERE month=? AND archived=FALSE AND family_id=?',
                                (cm, fid)).fetchall()
        if not payments:
            return jsonify({'error': 'אין תשלומים לארכוב'}), 400
        total = sum(p['amount'] for p in payments)
        count = len(payments)
        conn.execute('INSERT INTO archived_cycles (family_id,label,total,count,month) VALUES (?,?,?,?,?)',
                     (fid, label, total, count, cm))
        conn.execute('UPDATE payments SET archived=TRUE WHERE month=? AND archived=FALSE AND family_id=?', (cm, fid))

    user_id = request.api_user['user_id']
    user_name = request.api_user['display_name']
    send_push_to_family(fid,
        '📦 מחזור אורכב',
        f'{user_name} ארכב/ה את {label} — ₪{total:,.0f}',
        exclude_user_id=user_id)

    return jsonify({'message': f'{label} אורכב!', 'archived': {'label': label, 'total': total, 'count': count}})


# --- PAYMENTS: EXPORT CSV ---
@app.route('/api/payments/export', methods=['GET'])
@csrf.exempt
@require_api_family
def api_export_csv():
    fid = request.api_user['family_id']
    cm = get_cycle_month(fid)
    with get_db() as conn:
        payments = conn.execute(
            'SELECT description,amount,category,date FROM payments WHERE month=? AND archived=FALSE AND family_id=? ORDER BY date DESC',
            (cm, fid)).fetchall()
    output = io.StringIO()
    output.write('\ufeff')
    output.write('תיאור,סכום,קטגוריה,תאריך\n')
    for p in payments:
        output.write(f'{p["description"]},{p["amount"]},{p["category"]},{p["date"]}\n')
    return send_file(io.BytesIO(output.getvalue().encode('utf-8-sig')),
                     mimetype='text/csv', as_attachment=True, download_name=f'ourhome_{cm}.csv')


# ──────────────────────────────────────────────
# PUSH NOTIFICATIONS API
# ──────────────────────────────────────────────

# --- REGISTER PUSH TOKEN ---
@app.route('/api/push/register', methods=['POST'])
@csrf.exempt
@require_auth
def api_register_push_token():
    """Register a device push token for notifications"""
    data = request.get_json(silent=True) or {}
    token = data.get('token', '').strip()
    platform = data.get('platform', 'android')

    if not token:
        return jsonify({'error': 'טוקן נדרש'}), 400

    user_id = request.api_user['user_id'] if hasattr(request, 'api_user') else session['user_id']
    with get_db() as conn:
        existing = conn.execute('SELECT id, user_id FROM push_tokens WHERE token=?', (token,)).fetchone()
        if existing:
            conn.execute('UPDATE push_tokens SET user_id=?, platform=? WHERE token=?',
                         (user_id, platform, token))
        else:
            conn.execute('INSERT INTO push_tokens (user_id, token, platform) VALUES (?,?,?)',
                         (user_id, token, platform))

    return jsonify({'message': 'טוקן נרשם בהצלחה'})


# --- UNREGISTER PUSH TOKEN ---
@app.route('/api/push/unregister', methods=['POST'])
@csrf.exempt
@require_api_auth
def api_unregister_push_token():
    """Remove a push token (e.g. on logout)"""
    data = request.get_json(silent=True) or {}
    token = data.get('token', '').strip()

    if not token:
        return jsonify({'error': 'טוקן נדרש'}), 400

    with get_db() as conn:
        conn.execute('DELETE FROM push_tokens WHERE token=?', (token,))

    return jsonify({'message': 'טוקן הוסר'})


# --- SEND PUSH TO FAMILY (utility function — FCM V1 API) ---
FCM_SERVICE_ACCOUNT_PATH = os.environ.get(
    'FCM_SERVICE_ACCOUNT',
    os.path.join(os.path.dirname(os.path.abspath(__file__)), 'firebase-service-account.json')
)
_fcm_credentials = None


def get_fcm_access_token():
    """Get OAuth2 access token for FCM V1 API"""
    global _fcm_credentials
    try:
        from google.oauth2 import service_account
        import google.auth.transport.requests

        if not os.path.exists(FCM_SERVICE_ACCOUNT_PATH):
            print(f'FCM service account not found: {FCM_SERVICE_ACCOUNT_PATH}')
            return None

        if _fcm_credentials is None or not _fcm_credentials.valid:
            _fcm_credentials = service_account.Credentials.from_service_account_file(
                FCM_SERVICE_ACCOUNT_PATH,
                scopes=['https://www.googleapis.com/auth/firebase.messaging']
            )

        if not _fcm_credentials.valid:
            _fcm_credentials.refresh(google.auth.transport.requests.Request())

        return _fcm_credentials.token
    except Exception as e:
        print(f'FCM auth error: {e}')
        return None


def get_fcm_project_id():
    """Get Firebase project ID from service account file"""
    try:
        with open(FCM_SERVICE_ACCOUNT_PATH) as f:
            data = json.load(f)
            return data.get('project_id', '')
    except:
        return ''


def _send_push_worker(tokens_list, title, body, access_token, project_id):
    """Background worker that actually sends push notifications"""
    import urllib.request as urlreq
    url = f'https://fcm.googleapis.com/v1/projects/{project_id}/messages:send'

    for token_val in tokens_list:
        payload = json.dumps({
            'message': {
                'token': token_val,
                'notification': {'title': title, 'body': body},
                'android': {'notification': {'sound': 'default', 'icon': 'ic_notification'}},
                'data': {'title': title, 'body': body}
            }
        }).encode()

        req = urlreq.Request(url, data=payload, headers={
            'Content-Type': 'application/json',
            'Authorization': f'Bearer {access_token}'
        })
        try:
            urlreq.urlopen(req, timeout=10)
        except Exception as e:
            print(f'Push send error: {e}')


def send_push_to_family(family_id, title, body, exclude_user_id=None):
    """Send push notification to all family members (non-blocking, runs in background)"""
    try:
        with get_db() as conn:
            if exclude_user_id:
                tokens = conn.execute(
                    'SELECT token FROM push_tokens WHERE user_id IN '
                    '(SELECT id FROM users WHERE family_id=? AND id != ?)',
                    (family_id, exclude_user_id)
                ).fetchall()
            else:
                tokens = conn.execute(
                    'SELECT token FROM push_tokens WHERE user_id IN '
                    '(SELECT id FROM users WHERE family_id=?)',
                    (family_id,)
                ).fetchall()

        if not tokens:
            return

        access_token = get_fcm_access_token()
        if not access_token:
            print('FCM: No access token — push skipped')
            return

        project_id = get_fcm_project_id()
        if not project_id:
            print('FCM: No project ID — push skipped')
            return

        # Send in background thread so API response isn't delayed
        token_list = [t['token'] for t in tokens]
        t = threading.Thread(target=_send_push_worker,
                             args=(token_list, title, body, access_token, project_id),
                             daemon=True)
        t.start()

    except Exception as e:
        print(f'Push error: {e}')


# ──────────────────────────────────────────────
# FAMILY SETTINGS API
# ──────────────────────────────────────────────

@app.route('/api/family/settings', methods=['GET'])
@csrf.exempt
@require_auth
def api_get_family_settings():
    """Get family notification & budget settings"""
    fid = get_family_id()
    if not fid:
        return jsonify({'error': 'אין משפחה'}), 403
    with get_db() as conn:
        s = conn.execute('SELECT * FROM family_settings WHERE family_id=?', (fid,)).fetchone()
    if not s:
        return jsonify({
            'feeding_reminder_hours': 0,
            'budget_monthly': 0,
            'budget_daily': 0,
            'cycle_day': 1
        })
    return jsonify({
        'feeding_reminder_hours': s['feeding_reminder_hours'],
        'budget_monthly': s['budget_monthly'],
        'budget_daily': s['budget_daily'],
        'cycle_day': s['cycle_day'] if s['cycle_day'] else 1
    })


@app.route('/api/family/settings', methods=['PUT'])
@csrf.exempt
@require_auth
def api_update_family_settings():
    """Update family notification & budget settings"""
    fid = get_family_id()
    if not fid:
        return jsonify({'error': 'אין משפחה'}), 403
    data = request.get_json(silent=True) or {}

    new_cd = None
    with get_db() as conn:
        existing = conn.execute('SELECT id FROM family_settings WHERE family_id=?', (fid,)).fetchone()
        if existing:
            if 'feeding_reminder_hours' in data:
                conn.execute('UPDATE family_settings SET feeding_reminder_hours=? WHERE family_id=?',
                             (data['feeding_reminder_hours'], fid))
            if 'budget_monthly' in data:
                conn.execute('UPDATE family_settings SET budget_monthly=? WHERE family_id=?',
                             (data['budget_monthly'], fid))
            if 'budget_daily' in data:
                conn.execute('UPDATE family_settings SET budget_daily=? WHERE family_id=?',
                             (data['budget_daily'], fid))
            if 'cycle_day' in data:
                # Only family creator can change cycle day
                family = conn.execute('SELECT created_by FROM families WHERE id=?', (fid,)).fetchone()
                current_uid = int(request.api_user['user_id'] if hasattr(request, 'api_user') else session.get('user_id'))
                if not family or int(family['created_by']) != current_uid:
                    return jsonify({'error': 'רק מנהל המשפחה יכול לשנות את יום האיפוס'}), 403
                new_cd = int(data['cycle_day'])
                if new_cd < 1 or new_cd > 28:
                    return jsonify({'error': 'יום איפוס חייב להיות בין 1 ל-28'}), 400
                conn.execute('UPDATE family_settings SET cycle_day=? WHERE family_id=?', (new_cd, fid))
        else:
            new_cd = int(data.get('cycle_day', 1))
            conn.execute(
                'INSERT INTO family_settings (family_id, feeding_reminder_hours, budget_monthly, budget_daily, cycle_day) '
                'VALUES (?,?,?,?,?)',
                (fid,
                 data.get('feeding_reminder_hours', 0),
                 data.get('budget_monthly', 0),
                 data.get('budget_daily', 0),
                 new_cd))

    # Re-categorize non-archived payments in a separate connection
    if new_cd is not None:
        recat_conn = sqlite3.connect(DATABASE, timeout=10)
        recat_conn.row_factory = sqlite3.Row
        try:
            payments = recat_conn.execute(
                'SELECT id, date FROM payments WHERE family_id=? AND archived=0', (fid,)).fetchall()
            for p in payments:
                try:
                    date_str = p['date'].split('.')[0] if '.' in p['date'] else p['date']
                    pay_date = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
                    if new_cd == 1:
                        new_month = pay_date.strftime('%Y-%m')
                    elif pay_date.day >= new_cd:
                        new_month = pay_date.strftime('%Y-%m')
                    else:
                        prev = pay_date.replace(day=1) - timedelta(days=1)
                        new_month = prev.strftime('%Y-%m')
                    recat_conn.execute('UPDATE payments SET month=? WHERE id=?', (new_month, p['id']))
                except Exception as e:
                    print(f'Re-categorize error for payment {p["id"]}: {e}, date={p["date"]}')
            recat_conn.commit()
        finally:
            recat_conn.close()

    return jsonify({'message': 'הגדרות עודכנו!'})


@app.route('/api/family/cycle', methods=['GET'])
@csrf.exempt
@require_auth
def api_get_cycle_info():
    """Get current billing cycle info"""
    fid = get_family_id()
    if not fid:
        return jsonify({'error': 'אין משפחה'}), 403
    start, end, label = get_cycle_range(fid)
    cycle_day = get_cycle_day(fid)
    return jsonify({
        'cycle_day': cycle_day,
        'cycle_month': get_cycle_month(fid),
        'label': label,
        'start_date': start.strftime('%Y-%m-%d'),
        'end_date': end.strftime('%Y-%m-%d')
    })


# ──────────────────────────────────────────────
# BUDGET CHECK (called after adding payment)
# ──────────────────────────────────────────────
def check_budget_alerts(family_id):
    """Check if budget thresholds crossed and send alerts"""
    try:
        now = now_israel()
        cm = get_cycle_month(family_id)
        today = now.strftime('%Y-%m-%d')

        with get_db() as conn:
            settings = conn.execute('SELECT * FROM family_settings WHERE family_id=?', (family_id,)).fetchone()
            if not settings:
                print(f'Budget check: no settings for family {family_id}')
                return

            # Monthly budget check
            budget_monthly = settings['budget_monthly'] or 0
            if budget_monthly > 0:
                total = conn.execute(
                    'SELECT COALESCE(SUM(amount),0) as total FROM payments WHERE month=? AND archived=FALSE AND family_id=?',
                    (cm, family_id)).fetchone()['total']
                pct = (total / budget_monthly) * 100
                alert_100 = settings['budget_alert_100_sent'] or ''
                alert_80 = settings['budget_alert_80_sent'] or ''

                print(f'Budget check: family={family_id}, total={total}, budget={budget_monthly}, pct={pct:.0f}%, alert_80={alert_80}, alert_100={alert_100}')

                if pct >= 100 and alert_100 != cm:
                    send_push_to_family(family_id,
                        '🚨 חריגה מהתקציב החודשי!',
                        f'הוצאתם ₪{total:,.0f} מתוך ₪{budget_monthly:,} — חריגה!')
                    conn.execute('UPDATE family_settings SET budget_alert_100_sent=? WHERE family_id=?', (cm, family_id))
                    print(f'Budget alert 100% sent for family {family_id}')
                elif pct >= 80 and alert_80 != cm:
                    send_push_to_family(family_id,
                        '⚠️ התקציב החודשי עומד להיגמר',
                        f'הוצאתם ₪{total:,.0f} מתוך ₪{budget_monthly:,} ({pct:.0f}%)')
                    conn.execute('UPDATE family_settings SET budget_alert_80_sent=? WHERE family_id=?', (cm, family_id))
                    print(f'Budget alert 80% sent for family {family_id}')

            # Daily budget check
            budget_daily = settings['budget_daily'] or 0
            if budget_daily > 0:
                daily_total = conn.execute(
                    'SELECT COALESCE(SUM(amount),0) as total FROM payments WHERE date(date)=? AND archived=FALSE AND family_id=?',
                    (today, family_id)).fetchone()['total']
                if daily_total > budget_daily:
                    send_push_to_family(family_id,
                        '💸 חריגה מהתקציב היומי!',
                        f'הוצאתם היום ₪{daily_total:,.0f} מתוך ₪{budget_daily:,}')
                    print(f'Daily budget alert sent for family {family_id}')

    except Exception as e:
        print(f'Budget check error: {e}')
        import traceback
        traceback.print_exc()


# ──────────────────────────────────────────────
# AUTO-ARCHIVE SCHEDULER
# ──────────────────────────────────────────────
def check_auto_archive():
    """Auto-archive previous cycle when a new cycle starts — runs every 5 minutes"""
    while True:
        import time
        time.sleep(300)  # Check every 5 minutes

        try:
            now = now_israel()
            with get_db() as conn:
                families = conn.execute(
                    'SELECT fs.*, f.name as family_name FROM family_settings fs '
                    'JOIN families f ON fs.family_id = f.id'
                ).fetchall()

                for fam in families:
                    fid = fam['family_id']
                    cycle_day = fam['cycle_day'] or 1
                    last_archived = fam['last_cycle_archived'] or ''

                    # Calculate current cycle month
                    current_cm = get_cycle_month(fid)

                    # If we already archived this cycle, skip
                    if last_archived == current_cm:
                        continue

                    # Calculate previous cycle month
                    cm_year, cm_month = int(current_cm.split('-')[0]), int(current_cm.split('-')[1])
                    if cm_month == 1:
                        prev_cm = f'{cm_year - 1}-12'
                    else:
                        prev_cm = f'{cm_year}-{cm_month - 1:02d}'

                    # Check if previous cycle has unarchived payments
                    r = conn.execute(
                        'SELECT COALESCE(SUM(amount),0) as total, COUNT(*) as count '
                        'FROM payments WHERE month=? AND archived=FALSE AND family_id=?',
                        (prev_cm, fid)).fetchone()

                    if r['count'] == 0:
                        # No payments to archive, just mark as done
                        conn.execute('UPDATE family_settings SET last_cycle_archived=? WHERE family_id=?',
                                     (current_cm, fid))
                        continue

                    # Build label for the archived cycle
                    hebrew_months = {1:'ינואר',2:'פברואר',3:'מרץ',4:'אפריל',5:'מאי',6:'יוני',
                                     7:'יולי',8:'אוגוסט',9:'ספטמבר',10:'אוקטובר',11:'נובמבר',12:'דצמבר'}
                    prev_month_num = int(prev_cm.split('-')[1])
                    prev_year = int(prev_cm.split('-')[0])

                    if cycle_day == 1:
                        label = f'{hebrew_months[prev_month_num]} {prev_year}'
                    else:
                        end_day = cycle_day - 1
                        next_m = prev_month_num + 1 if prev_month_num < 12 else 1
                        label = f'{hebrew_months[prev_month_num]} {prev_year} ({cycle_day}.{prev_month_num}–{end_day}.{next_m})'

                    # Archive!
                    conn.execute('INSERT INTO archived_cycles (family_id,label,total,count,month) VALUES (?,?,?,?,?)',
                                 (fid, label, r['total'], r['count'], prev_cm))
                    conn.execute('INSERT OR REPLACE INTO app_settings (key,value) VALUES (?,?)',
                                 (f'last_archived_total_{fid}', str(r['total'])))
                    conn.execute('UPDATE payments SET archived=TRUE WHERE month=? AND archived=FALSE AND family_id=?',
                                 (prev_cm, fid))
                    conn.execute('UPDATE family_settings SET last_cycle_archived=? WHERE family_id=?',
                                 (current_cm, fid))

                    print(f'Auto-archive: family {fid} ({fam["family_name"]}): {label} — {r["count"]} payments, ₪{r["total"]:.0f}')

                    # Push notification to family
                    send_push_to_family(fid,
                        '📊 מחזור חדש התחיל',
                        f'סה"כ {label}: ₪{r["total"]:,.0f} ({r["count"]} תשלומים)')

        except Exception as e:
            print(f'Auto-archive error: {e}')
            import traceback
            traceback.print_exc()


# Start auto-archive in background
_archive_thread = threading.Thread(target=check_auto_archive, daemon=True)
_archive_thread.start()


# ──────────────────────────────────────────────
# FEEDING REMINDER SCHEDULER
# ──────────────────────────────────────────────
def check_feeding_reminders():
    """Check all families for feeding reminders — runs every 60 seconds"""
    while True:
        try:
            now = now_israel()
            today = now.strftime('%Y-%m-%d')

            with get_db() as conn:
                families = conn.execute(
                    'SELECT fs.*, f.name as family_name FROM family_settings fs '
                    'JOIN families f ON fs.family_id = f.id '
                    'WHERE fs.feeding_reminder_hours > 0'
                ).fetchall()

                for fam in families:
                    fid = fam['family_id']
                    hours = fam['feeding_reminder_hours']

                    # Find last feeding (bottle, breastfeeding, or solid)
                    last = conn.execute(
                        'SELECT date FROM feedings WHERE family_id=? AND feeding_type IN (?,?,?) '
                        'ORDER BY date DESC LIMIT 1',
                        (fid, 'bottle', 'breastfeeding', 'solid')
                    ).fetchone()

                    if not last:
                        continue

                    try:
                        last_dt = datetime.strptime(last['date'].split('.')[0], '%Y-%m-%d %H:%M:%S')
                        diff_minutes = (now - last_dt).total_seconds() / 60
                        threshold_minutes = hours * 60

                        # Check if we should alert
                        if diff_minutes >= threshold_minutes:
                            # Don't alert again if we already did for this feeding
                            last_alert = fam['last_feeding_alert']
                            if last_alert:
                                try:
                                    alert_dt = datetime.strptime(last_alert.split('.')[0], '%Y-%m-%d %H:%M:%S')
                                    # Skip if last alert was after the last feeding
                                    if alert_dt > last_dt:
                                        continue
                                except:
                                    pass

                            hours_passed = diff_minutes / 60
                            send_push_to_family(fid,
                                '🍼 תזכורת האכלה',
                                f'עברו {hours_passed:.1f} שעות מהאכלה אחרונה')
                            conn.execute(
                                'UPDATE family_settings SET last_feeding_alert=? WHERE family_id=?',
                                (now.strftime('%Y-%m-%d %H:%M:%S'), fid))

                    except Exception as e:
                        print(f'Feeding reminder parse error: {e}')

        except Exception as e:
            print(f'Feeding reminder error: {e}')

        import time
        time.sleep(60)  # Check every 60 seconds for precision


# Start feeding reminder in background
_reminder_thread = threading.Thread(target=check_feeding_reminders, daemon=True)
_reminder_thread.start()


init_db()

if __name__ == '__main__':
    init_db()
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)